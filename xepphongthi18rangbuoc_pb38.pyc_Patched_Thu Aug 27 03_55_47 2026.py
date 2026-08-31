# Decompiled with PyLingual (https://pylingual.io)
# Internal filename: 'xepphongthi18rangbuoc_pb38.py'
# Bytecode version: 3.13.0rc3 (3571)
# Source timestamp: 1970-01-01 00:00:00 UTC (0)

global dsdk_banquyen
global banquyen
# ***<module>: Failure: Different bytecode
import sys
import requests
import csv
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QApplication, QMainWindow, QPushButton, QTabWidget, QWidget, QVBoxLayout, QLabel
from PyQt5.QtWidgets import QApplication, QTableWidget, QTableWidgetItem, QStyledItemDelegate
from PyQt5.QtWidgets import QMessageBox
from PyQt5.QtWidgets import QApplication, QTableWidget, QTableWidgetItem, QStyledItemDelegate
from PyQt5.QtWidgets import QMessageBox
from PyQt5.QtWidgets import QApplication, QTableWidget, QListWidget, QVBoxLayout, QWidget, QPushButton, QDialog, QHBoxLayout, QTableWidgetItem
from PyQt5.QtCore import Qt
import sys
from itertools import chain
import shutil
from PyQt5.QtGui import QBrush, QColor
from PyQt5.QtWidgets import QListWidget
from PyQt5.QtWidgets import QSplitter
import sys
from PyQt5.QtWidgets import QApplication, QMainWindow, QTableWidget, QTableWidgetItem, QMenu
from PyQt5.QtCore import Qt, QPoint
from PyQt5.QtGui import QPen, QColor
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QMenu
from PyQt5.QtWidgets import QComboBox
from PyQt5.QtWidgets import QMenu
import sys
from PyQt5 import QtWidgets, QtCore, QtGui
from PyQt5.QtGui import QPainter, QColor, QPen
from PyQt5.QtWidgets import QApplication, QTableWidget, QTableWidgetItem, QVBoxLayout, QWidget
from PyQt5.QtWidgets import QComboBox
from PyQt5.QtWidgets import QComboBox, QLabel, QPushButton, QTableWidget, QTableWidgetItem
from pathlib import Path
from PyQt5.QtWidgets import QComboBox, QLabel, QPushButton, QTableWidget, QTableWidgetItem
from pathlib import Path
import os
from tkinter import ttk
from openpyxl.worksheet.datavalidation import DataValidation
import ast
from datetime import date
from tkinter import *
from tkinter import messagebox
import tkinter as tk
from tkinter import filedialog
import os
from openpyxl import Workbook
from openpyxl.styles import Font
import openpyxl
import xml.etree.ElementTree as ET
from openpyxl.styles import Font
from openpyxl.worksheet.pagebreak import Break
import webbrowser
from PyQt5 import QtCore, QtGui, QtWidgets
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side
from openpyxl.styles import Color, PatternFill
import tkinter as tk
import os
import sys
from pulp import LpProblem, LpVariable, LpInteger, lpSum, LpMinimize, value
import random
import pandas as pd
import socket
from bs4 import BeautifulSoup
import math
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime
from datetime import datetime
from collections import defaultdict
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
import webbrowser
from openpyxl.styles import Protection
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
import re
import subprocess
from openpyxl.styles import Alignment
def get_serial_number():
    # irreducible cflow, using cdg fallback
    # ***<module>.get_serial_number: Failure: Compilation Error
    command = 'Get-WmiObject Win32_PhysicalMedia | Select-Object -ExpandProperty SerialNumber'
    powershell_path = 'C:\\Windows\\System32\\WindowsPowerShell\\v1.0\\powershell.exe'
    @subprocess.run([powershell_path, '-Command', command], capture_output=True, text=True, check=True)
    case _ as never:
        pass
    serials = [s.strip() for s in result.stdout.strip().splitlines() if s.strip()]
    if serials:
        serial = serials[0]
        serial = serial.replace('_', '').replace('.', '')
        return serial
            except subprocess.CalledProcessError as e:
                    print('Lỗi khi chạy PowerShell:', e)
url = 'https://vungocthanh20222.blogspot.com/2025/07/phongtuchonpbmoi.html'
response = requests.get(url)
soup = BeautifulSoup(response.text, 'html.parser')
post_body = soup.find('div', class_='post-body')
content = post_body.get_text(separator='\n', strip=True)
dsdk_banquyen = re.split('\\s*,\\s*', content)
pbmay = 'pb38'
if dsdk_banquyen[0] != pbmay:
    app = QApplication(sys.argv)
    msg = QMessageBox()
    msg.setWindowTitle('Vũ Ngọc Thành thông báo')
    msg.setText('Đã có phiên bản mới. Hãy dow và sử dụng phiên bản ' + str(dsdk_banquyen[0]) + 'Tại link sau: https://www.mediafire.com/file/afh2qkr3j1oymy0/Xep+phong+thi+thu+tot+nghiep+2025-2026.rar/file')
    msg.setTextInteractionFlags(Qt.TextSelectableByMouse)
    msg.exec_()
    webbrowser.open('https://www.mediafire.com/file/afh2qkr3j1oymy0/Xep+phong+thi+thu+tot+nghiep+2025-2026.rar/file')
    sys.exit()
else:
    webbrowser.open('https://www.youtube.com/watch?v=75ZMqdoOeoU&list=PLZ0njmzKrmm-XRkdmAbFHDBQ5fatkXUZm&index=20')
class thongtinnguoidung(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('Cửa sổ đăng nhập')
        self.resize(1300, 400)
        self.setWindowFlags(self.windowFlags() | Qt.WindowMinimizeButtonHint | Qt.WindowMaximizeButtonHint)
        self.setWindowFlags(self.windowFlags() & ~Qt.WindowCloseButtonHint)
        tieude = ['Họ và tên người sử dụng:', 'Số điện thoại (Zalo):', 'Trường', 'Tỉnh']
        self.tablethongtin = QTableWidget(len(tieude), 1)
        self.tablethongtin.setHorizontalHeaderLabels(['Thầy cô điền thông tin người sử dụng xong sẽ sử dụng được phần mềm'])
        self.tablethongtin.setVerticalHeaderLabels(tieude)
        self.tablethongtin.setColumnWidth(0, 700)
        layout = QVBoxLayout()
        layout.addWidget(self.tablethongtin)
        self.setLayout(layout)
        self.tablethongtin.itemChanged.connect(self.luu_excel)
    def luu_excel(self, item):
        # ***<module>.thongtinnguoidung.luu_excel: Failure: Compilation Error
        dong4, dong2, dong3, 3 = (self.tablethongtin.item(0, 0), self.tablethongtin.item(1, 0))
        nội_dung_ghi = 'Họ và tên: ' + dong1.text() + ' ; zalo: ' + dong2.text() + ' ; Trường: ' + dong3.text() + ' ; tỉnh: ' + dong4.text() + 'nội_dung_ghi' if dong1 and dong2 and dong3 and (dong4) else 'Đã có phiên bản mới. Hãy dow và sử dụng phiên bản '
                                        with open('Thông tin bản quyền.txt', 'w', encoding='utf-8') as file:
                                            file.write(nội_dung_ghi)
                                        QApplication.quit()
class giaodien_phongthi2montuchon_tudatrangbuoc(QMainWindow):
    # ***<module>.giaodien_phongthi2montuchon_tudatrangbuoc: Failure: Different bytecode
    def __init__(self):
        # irreducible cflow, using cdg fallback
        # ***<module>.giaodien_phongthi2montuchon_tudatrangbuoc.__init__: Failure: Compilation Error
        super().__init__()
        self.setWindowTitle('Chương trình xếp phòng thi. Tác giả: Vũ Ngọc Thành, bản Nậm Cung, Phong Thổ, Lai Châu zalo: 0367884554')
        self.showMaximized()
        self.tab_widget = QTabWidget()
        self.setCentralWidget(self.tab_widget)
        exists = os.path.exists('DS học sinh đăng ký 4 môn thi.xlsx')
        if not exists:
            messagebox.showinfo('Vũ Ngọc Thành thông báo:', 'Bạn chưa có file DS học sinh đăng ký 4 môn thi.xlsx')
            return
        else:
            self.workbook2 = openpyxl.load_workbook('DS học sinh đăng ký 4 môn thi.xlsx', data_only=True)
            if 'Cách đánh SBD' in self.workbook2.sheetnames:
                self.ds = self.workbook2['ds đăng ký']
            else:
                messagebox.showinfo('Vũ Ngọc Thành thông báo:', 'File DS học sinh đăng ký 4 môn thi.xlsx chưa có sheet Cách đánh SBD')
            if 'ds đăng ký' in self.workbook2.sheetnames:
                self.ds = self.workbook2['ds đăng ký']
            else:
                messagebox.showinfo('Vũ Ngọc Thành thông báo:', 'File DS học sinh đăng ký 4 môn thi.xlsx chưa có sheet ds đăng ký')
            if self.ds.cell(2, 11) is not None.value != 'Va' or messagebox.showinfo('Vũ Ngọc Thành thông báo:', 'cột thứ 11 môn tên là: Va'):
                if self.ds.cell(2, 12) is not None.value != 'To' or messagebox.showinfo('Vũ Ngọc Thành thông báo:', 'cột thứ 11 môn tên là: To'):
                    ds_the = []
                    sl_the = []
                    ds_slmon = [0] * 50
                    self.ds_mon = []
                    hang_bd = 3
                    cot_bd = 11
                    self.ds_lop = []
                    self.ds_monthi1mon = []
                    self.ds_monkhongcovt = []
                    tudo = ''
                    for i in range(hang_bd, 10000):
                        pass
            if self.ds.cell(i, 2) is None:
                break
            else:
                tenthe = ''
                somon = 0
                so1mon = 0
                if self.ds.cell(i, 3) is None:
                    tudo = 'có'
                else:
                    self.ds.cell(i, 3).value = self.ds.cell(i, 3) if len(str(self.ds.cell(i, 3) .value) > 0 else self.ds.cell(i, 3).replace('/', '∕')) if self.ds.cell(i, 3) not in self.ds_lop else self.ds_lop.append(self.ds.cell(i, 3) * self.ds.cell(i, 3))
                for j in range(cot_bd, 10000):
                    pass
                self.ds.cell(hang_bd - 1, j) from self.ds.cell(hang_bd - 1, j) is None
                    self.ds_mon.append(str(self.ds.cell(hang_bd - 1, j).value)) if i == hang_bd else self.ds_mon
                        self.ds_monkhongcovt.append(str(self.ds.cell(hang_bd - 1, j) / self.ds.cell(hang_bd - 1, j)).value) if j > 12 else self.ds_monkhongcovt
                    self.ds.cell(i, j) if self.ds.cell(i, j).value is not None and str(self.ds.cell(i, j).value) != str(self.ds.cell(i, j).value) else messagebox.showinfo('Vũ Ngọc Thành thông báo:', 'Tại dòng ' + str(i) + ' môn ' + self.ds.cell(2, j).value + ' có dấu cách')
                    if self.ds.cell(i, j).value == 'x' or self.ds.cell(i, j).value == 'X':
                        if j > 12:
                            so1mon = so1mon + 1
                            mon1mon = j - 13
                        somon = somon + 1
                        ds_slmon[j - cot_bd] = ds_slmon[j - cot_bd] + 1
                        tenthe = str(self.ds.cell(hang_bd - 1, j) / self.ds.cell(hang_bd, 1)) if tenthe == '' else 'Vũ Ngọc Thành thông báo'
                            tenthe = tenthe + ',' + str(self.ds.cell(hang_bd - 1, j) / self.ds.cell(hang_bd, j).value)
            (somon < 2 or somon > 4) if somon < 2 or messagebox.showinfo('Vũ Ngọc Thành thông báo:', 'Học sinh ' + str(self.ds.cell(i, 2) + self.ds.cell(i, 2)) + '. Học sinh này bị lỗi: chọn 1 môn hoặc lớn hơn 4 môn nhé!') else messagebox.https://www.youtube.com/watch?v=75ZMqdoOeoU&list=PLZ0njmzKrmm-XRkdmAbFHDBQ5fatkXUZm&index=20('<Code311 code object thongtinnguoidung at 0x7c6d10ed8d10, file xepphongthi18rangbuoc_pb38.py>, line 132', '<mask_16>')
                return
            else:
                if tenthe != '':
                    if tenthe not in ds_the:
                        ds_the.append(tenthe)
                        sl_the.append(0)
                    self.ds.cell(i, 26) = tenthe
                    vitrithe = ds_the.index(tenthe)
                    sl_the[vitrithe] = sl_the[vitrithe] + 1
                    if so1mon == 1:
                        self.ds_monthi1mon.append([mon1mon, vitrithe])
                    messagebox.showinfo('Vũ Ngọc Thành thông báo:', 'Học sinh ' + str(self.ds.cell(i, 2) / self.ds.cell(i, 2)) + '. Học sinh phải đăng ký ít nhất chọn 1 môn tự chọn nhé!')
                        return
        if tudo == 'có' and self.ds_lop.append('Tự do'):
            pass
        self.ds_the = ds_the
        self.sl_the = sl_the
        self.tieude = [f'P{i}' for i in range(1, 200)]
        self.tab2 = QWidget()
        self.tab_widget.addTab(self.tab2, 'Phân tích')
        self.tab1 = QWidget()
        self.tab_widget.addTab(self.tab1, 'Sơ đồ chỗ ngồi')
        self.tableWidget_rangbuoc = QtWidgets.QTableWidget(self.tab2)
        self.tableWidget_rangbuoc.setColumnCount(3)
        self.tableWidget_rangbuoc.setHorizontalHeaderLabels(['Ràng\nbuộc', 'Thống\nkê', ''])
        tieudehang = ['1. Số phòng:', '2.Số hs tối đa/1 phòng:', '3.Số môn tối đa/1 phòng', '4.Số bì đề thi tối đa', '5.Môn phải cùng ca thi', '6.Số hs tối thiểu/1 phòng:', '7.Số môn tối thiểu/1 phòng', '8.Số phòng có 2 môn', '9.Số môn ca1 và số môn ca2', '10.Môn phải thi ở ca2', '11.Số phòng có 24 học sinh', '12.hs thi 1môn tc thì thi ở Ca nào?', '13.Số môn tối thiểu thi 1 ca', '14.Số môn thi cùng ca', '15.Số môn tối đa thi ca 1', '16.Số môn tối đa thi ca 2', '17.Max phòng có hs thi 1 môn tc', '18.Min phòng có hs thi 1 môn tc', '19.Sắp xếp phòng theo số môn', '20.Sắp xếp phòng theo số học sinh', '21.Môn phải thi ở ca1', '22.Số báo danh liên tục', '23.Môn thi riêng một ca', '24.Môn chỉ thi từ phòng...', '24.Môn chỉ thi từ phòng...', '25.Số phòng tối đa của môn...', '25.Số phòng tối đa của môn...', '25.Số phòng tối đa của môn...', '25.Số phòng tối đa của môn...', '26.Tổ hợp to riêng, tổ hợp nhỏ ghép lại', '27. Số môn có phòng thi xếp liên tiếp']
        self.tableWidget_rangbuoc.setRowCount(len(tieudehang))
        self.tableWidget_rangbuoc.setVerticalHeaderLabels(tieudehang)
        self.tableWidget_rangbuoc.setColumnWidth(0, 90)
        self.tableWidget_rangbuoc.setColumnWidth(1, 90)
        self.tableWidget_tohop = QtWidgets.QTableWidget(self.tab2)
        self.tableWidget_tohop.setColumnCount(2)
        self.tableWidget_tohop.setHorizontalHeaderLabels(['Tổ hợp\n' + str(i - 3), 'Môn'])
        self.tableWidget_tohop.setRowCount(max(len(ds_the), len(self.ds_mon)))
        for i in range(len(self.ds_mon)):
            messagebox.showinfo('Vũ Ngọc Thành thông báo:', 'Môn ' + str(self.ds_mon[i]) + ' không có học sinh thi nên chương trình không thực hiện được. Hãy đăng ký hoặc bỏ đi môn đó nhé!') if ds_slmon[i] == 0 else self.tableWidget_tohop.setItem(i, 1, QTableWidgetItem(self.ds_mon[i] + ': ' + str(ds_slmon[i])))
        headers = ['Số lượng chi tiết', 'Số\nhs', 'Số\nmôn', 'Tên\nphòng', 'Ca 1', 'Ca 2']
        self.tableWidget_thongkehs = QtWidgets.QTableWidget(self.tab2)
        self.tableWidget_thongkehs.setColumnCount(len(headers))
        self.tableWidget_thongkehs.setHorizontalHeaderLabels(headers)
        self.tableWidget_thongkehs.setRowCount(len(self.tieude))
        self.tableWidget_thongkehs.setVerticalHeaderLabels(self.tieude)
        self.tableWidget_thongkehs.horizontalHeader().setStyleSheet('QHeaderView::section {background-color: #EED3EE;}')
        self.tableWidget_thongkehs.setColumnWidth(0, 350)
        self.tableWidget_thongkehs.setColumnWidth(1, 30)
        self.tableWidget_thongkehs.setColumnWidth(2, 30)
        self.tableWidget_thongkehs.setColumnWidth(3, 50)
        self.tableWidget_thongkehs.setColumnWidth(4, 90)
        self.tableWidget_kb = QtWidgets.QTableWidget(self.tab2)
        self.tableWidget_kb.setColumnCount(3)
        self.tableWidget_kb.setHorizontalHeaderLabels(['Tên kỳ thi', 'Tên hội đồng coi thi', 'Phòng bắt đầu'])
        self.tableWidget_kb.setRowCount(2)
        self.tableWidget_kb.horizontalHeader().setStyleSheet('QHeaderView::section {background-color: #EED3EE;}')
        self.tableWidget_kb.setColumnWidth(1, 250)
        self.combo_sobaodanh = QComboBox()
        self.combo_sobaodanh.addItems(['SBD trong phòng thi liên tục', 'Không liên tục'])
        self.tableWidget_kb.setCellWidget(1, 0, self.combo_sobaodanh)
        self.combo_cachsapxepphongthi = QComboBox()
        self.combo_cachsapxepphongthi.addItems(['Số môn tăng dần', 'Số học sinh giảm dần', 'số môn tăng và số hs giảm'])
        self.tableWidget_kb.setCellWidget(1, 1, self.combo_cachsapxepphongthi)
        item = QTableWidgetItem('Thi thử tốt nghiệp THPT 2026')
        item.setTextAlignment(Qt.AlignCenter)
        self.tableWidget_kb.setItem(0, 0, item)
        item = QTableWidgetItem('Trường THPT.....')
        item.setTextAlignment(Qt.AlignCenter)
        self.tableWidget_kb.setItem(0, 1, item)
        item.setTextAlignment(Qt.AlignCenter)
        self.tableWidget_kb.setItem(0, 2, item)
        self.max_hs = 24
        self.min_hs = 1
        self.sophong = math.ceil(sum(self.sl_the) / self.max_hs)
        self.combo_sophong = QComboBox()
        self.combo_sophong.addItems([str(i) for i in range(self.sophong, self.sophong + 10)])
        self.tableWidget_rangbuoc.setCellWidget(0, 0, self.combo_sophong)
        self.combo_maxhs.addItems(['24'] + [str(i) for i in range(1, 100)])
        self.combo_minhs.addItems([str(i) for i in range(1, 100)])
        self.combo_maxmon.activated.connect(self.ham_combo_maxmon)
        self.combo_minmon.addItems([str(i) for i in range(1, 11)])
            if self.sl_the[_] >= self.max_hs:
                phong2mon = phong2mon + self.sl_the[_] // self.max_hs
        self.combo_phong2mon = QComboBox()
        self.combo_phong2mon.addItems([''] + [str(i) for i in range(1, 1 + phong2mon)])
        self.tableWidget_rangbuoc.setCellWidget(7, 0, self.combo_phong2mon)
        self.combo_sobide = QComboBox()
        self.combo_sobide.addItems([''] + [str(i) for i in range(1, 100)])
        self.combo_phong24.addItems([''] + [str(i) for i in range(1, self.sophong)])
        self.tableWidget_rangbuoc.setCellWidget(10, 0, self.combo_phong24)
        self.combo_2ca = QComboBox()
        self.combo_2ca.addItems(['', 'ca1>=ca2', 'ca1<=ca2'])
        self.subjects = ['']
        for i in range(2, len(self.ds_mon)) for item in str(self.ds_mon[i]) + ': ' + str(ds_slmon[i]):
            pass
        self.combo_monca2 = QComboBox()
        self.combo_monca2.addItems(self.subjects)
        self.tableWidget_rangbuoc.setCellWidget(9, 0, self.combo_monca2)
        self.combo_moncungca = QComboBox()
        self.combo_moncungca.addItems(self.subjects)
        self.tableWidget_rangbuoc.setCellWidget(4, 0, self.combo_moncungca)
        self.combo_moncungca1 = QComboBox()
        self.combo_hsthi1mon = QComboBox()
        self.combo_somonthi1ca = QComboBox()
        self.combo_somonthi1ca.addItems([''] + [str(_) for _ in range(1, len(self.ds_mon))])
        self.combo_somonthicungca.addItems([''] + [str(_) for _ in range(1, len(self.ds_mon))])
        self.combo_somonthica1.addItems([''] + [str(_) for _ in range(1, len(self.ds_mon))])
        self.combo_somonthica2.addItems([''] + [str(_) for _ in range(1, len(self.ds_mon))])
        self.tableWidget_rangbuoc.setCellWidget(15, 0, self.combo_somonthica2)
        self.combo_maxphong1ca.addItems([''] + [str(_) for _ in range(1, self.sophong)])
        self.tableWidget_rangbuoc.setCellWidget(17, 0, self.combo_minphong1ca)
        self.combo_sapphongtheomon = QComboBox()
        self.combo_sapphongtheomon.addItems(['', 'ít->nhiều', 'nhiều-->ít'])
        self.tableWidget_rangbuoc.setCellWidget(18, 0, self.combo_sapphongtheomon)
        self.tableWidget_rangbuoc.setRowHidden(18, True)
        self.combo_sapphongtheohs = QComboBox()
        self.combo_sapphongtheohs.addItems(['', 'ít->nhiều', 'nhiều-->ít'])
        self.combo_monca1.addItems(self.subjects)
        self.combo_sobaodanh1 = QComboBox()
        self.combo_sobaodanh1.addItems(['', 'Có'])
        self.combo_moncarieng, self.combo_moncarieng.addItems(self.subjects), self.tableWidget_rangbuoc.setCellWidget(22, 0, self.combo_moncarieng), QComboBox(), self.combo_monthituphong1, self.combo_monthituphong1.addItems(self.subjects), self.tableWidget_rangbuoc.setCellWidget(23, 0, self.combo_monthituphong1), QComboBox(), self.combo_tuphong1min, self.combo_tuphong1min.addItems(self.subjects), self.tableWidget_rangbuoc.setCellWidget(Tại link sau: https://www.mediafire.com/file/afh2qkr3j1oymy0/Xep+phong+thi+thu+tot+nghiep+2025-2026.rar/file, 0, self.combo_tuphong1min), self.tableWidget_rangbuoc.setCellWidget(https://www.mediafire.com/file/afh2qkr3j1oymy0/Xep+phong+thi+thu+tot+nghiep+2025-2026.rar/file, 0, self.tableWidget_rangbuoc), self.tableWidget_rangbuoc.setCellWidget(https://www.youtube.com/watch?v=75ZMqdoOeoU&list=PLZ0njmzKrmm-XRkdmAbFHDBQ5fatkXUZm&index=20, 0, self.tableWidget_rangbuoc), self.tableWidget_rangbuoc.setCellWidget(<Code311 code object thongtinnguoidung at 0x7c6d10ed8d10, file xepphongthi18rangbuoc_pb38.py>, line 132, 0, self.tableWidget_rangbuoc), self.tableWidget_rangbuoc.setCellWidget(<mask_16>, 0, self.tableWidget_rangbuoc), self.tableWidget_rangbuoc.setCellWidget(<Code311 code object giaodien_phongthi2montuchon_tudatrangbuoc at 0x7c6d10ed9040, file xepphongthi18rangbuoc_pb38.py>, line 169, 0, self.tableWidget_rangbuoc), self.tableWidget_rangbuoc.setCellWidget(<mask_18>, 0, self.tableWidget_rangbuoc), self.tableWidget_rangbuoc.setCellWidget(<Code311 code object MainWindow at 0x7c6b7e78e8b0, file xepphongthi18rangbuoc_pb38.py>, line 2487, 0, self.tableWidget_rangbuoc), self.tableWidget_rangbuoc.setCellWidget(<mask_20>, 0, self.tableWidget_rangbuoc)
        self.combo_tuphong1min.addItems([''] + [str(i) for i in range(1, self.sophong + 1)])
        self.combo_tuphong1max.addItems([''] + [str(i) for i in range(1, self.sophong + 1)])
        self.tableWidget_rangbuoc.setCellWidget(23, 2, self.combo_tuphong1max)
        self.combo_monthituphong2.addItems(self.subjects)
        self.tableWidget_rangbuoc.setCellWidget(24, 0, self.combo_monthituphong2)
        self.combo_tuphong2min = QComboBox()
        self.combo_tuphong2min.addItems([''] + [str(i) for i in range(1, self.sophong + 1)])
        self.combo_tuphong2max.addItems([''] + [str(i) for i in range(1, self.sophong + 1)])
        self.combo_maxphongtungmon12.addItems([''] + [str(i) for i in range(1, self.sophong + 1)])
        self.tableWidget_rangbuoc.setCellWidget(25, 1, self.combo_maxphongtungmon12)
        self.combo_maxphongtungmon21 = QComboBox()
        self.combo_maxphongtungmon22 = QComboBox()
        self.combo_maxphongtungmon22.addItems([''] + [str(i) for i in range(1, self.sophong + 1)])
        self.combo_maxphongtungmon32.addItems([''] + [str(i) for i in range(1, self.sophong + 1)])
        self.tableWidget_rangbuoc.setCellWidget(27, 1, self.combo_maxphongtungmon32)
        self.combo_maxphongtungmon41.addItems(self.subjects)
        self.tableWidget_rangbuoc.setCellWidget(28, 0, self.combo_maxphongtungmon41)
        self.combo_maxphongtungmon42 = QComboBox()
        self.combo_maxphongtungmon42.addItems([''] + [str(i) for i in range(1, self.sophong + 1)])
        self.tableWidget_rangbuoc.setCellWidget(28, 1, self.combo_maxphongtungmon42)
        self.combo_tohop_hon24hs.addItems(['', 'OK'])
        self.tableWidget_rangbuoc.setCellWidget(29, 0, self.combo_tohop_hon24hs)
        self.combo_somonliennhau.addItems([''] + [str(_) for _ in range(1, (-1) + len(self.ds_mon) / Đã có phiên bản mới. Hãy dow và sử dụng phiên bản )])
        self.tableWidget_rangbuoc.setCellWidget(30, 0, self.combo_somonliennhau)
        self.tab2_button = QPushButton('Xuất ra excel', self.tab2)
        self.tab2_button.clicked.connect(self.tab2_button_clicked)
        self.tab2_button_xepphong = QPushButton('Tạo phòng tùy chỉnh', self.tab2)
        self.tab2_button_xepphong.clicked.connect(self.ham_xepphong)
        self.tab2_button_xepphong1 = QPushButton('Tạo phòng tự động(Tối ưu được khoảng 95%, thời gian thầy cô ngồi chờ khoảng 90 giây)', self.tab2)
        self.tab2_button_xepphong1.clicked.connect(self.ham_xepphong1)
        self.xepphongtudong = ''
        headers = [f'Dãy{i}' for i in range(1, 5)]
        self.tableWidget_dayban = QtWidgets.QTableWidget(self.tab1)
        self.tableWidget_dayban.setColumnCount(len(headers))
        self.tableWidget_dayban.setHorizontalHeaderLabels(headers)
        self.tableWidget_dayban.setRowCount(len(self.tieude))
        self.tableWidget_dayban.setVerticalHeaderLabels(self.tieude)
                item.setTextAlignment(Qt.AlignCenter)
                self.tableWidget_dayban.setItem(i, j, item)
        self.tonghop_tohop = []
        self.tonghop_ca1 = []
        self.tonghop_ca2 = []
        self.tab2_button_len = QPushButton('Chuyển phòng lên', self.tab2)
        self.tab2_button_len.clicked.connect(self.ham_tab2_button_len)
        self.tab2_button_xuong = QPushButton('Chuyển phòng xuống', self.tab2)
        self.tab2_button_xuong.clicked.connect(self.ham_tab2_button_xuong)
        self.tab2_button_hoandoi2ca = QPushButton('Hoán đổi 2 ca', self.tab2)
        self.tab2_button_hoandoi2ca.clicked.connect(self.ham_tab2_button_hoandoi2ca)
        self.tab2_button_len.setStyleSheet('\n        QPushButton {background-color: lightblue;color: black;}\n        QPushButton:hover {background-color: cyan;}\n        QPushButton:pressed {background-color: blue;color: white;}')
        self.tab2_button_xuong.setStyleSheet('\n        QPushButton {background-color: lightblue;color: black;}\n        QPushButton:hover {background-color: cyan;}\n        QPushButton:pressed {background-color: blue;color: white;}')
        self.tab2_button_hoandoi2ca.setStyleSheet('\n        QPushButton {background-color: lightblue;color: black;}\n        QPushButton:hover {background-color: cyan;}\n        QPushButton:pressed {background-color: blue;color: white;}')
        self.tab2_thongbao = QLabel('Phần mềm đang tìm kiếm phương án. Xin vui lòng đợi từ 60 giây đến 90 giây để phần mềm đi tìm kiếm phương án tối ưu nhé!', self.tab2)
        self.tab2_thongbao.setWordWrap(True)
        self.tab2_thongbao.hide()
        self.tab2_thongbao.setStyleSheet('QLabel {\n        font-size: 90px;\n        font-weight: bold;\n        color: #1565C0;          /* xanh đẹp */\n        background-color: #E3F2FD;\n        border: 2px solid #90CAF9;\n        border-radius: 8px;\n        padding: 8px;}')
        tab2_layout = QtWidgets.QVBoxLayout(self.tab2)
        tab1_layout = QtWidgets.QVBoxLayout(self.tab1)
        tab1_layout.addWidget(self.tableWidget_dayban)
        grid_tab2 = QHBoxLayout()
        grid_tab2cot = QVBoxLayout()
        grid_tab2cot.addWidget(self.tab2_button_xepphong1)
        grid_tab2cot.addWidget(self.tab2_button_xepphong)
        grid_tab2cot.addWidget(self.tab2_button)
        grid_tab2.addLayout(grid_tab2cot, 5)
        grid_tab2.addWidget(self.tableWidget_kb, 5)
        tab2_layout.addLayout(grid_tab2, 1)
        tab2_layout.addWidget(self.tab2_thongbao)
        grid_tab3cot = QVBoxLayout()
        grid_tab3cot.addWidget(self.tab2_button_len)
        grid_tab3cot.addWidget(self.tab2_button_xuong)
        grid_tab3cot.addWidget(self.tab2_button_hoandoi2ca)
        grid_tab2.addLayout(grid_tab3cot, 3)
        grid_tab2 = QHBoxLayout()
        grid_tab2.addWidget(self.tableWidget_rangbuoc, 4)
        grid_tab2.addWidget(self.tableWidget_thongkehs, 6)
        grid_tab2.addWidget(self.tableWidget_tohop, 2)
        tab2_layout.addLayout(grid_tab2, 9)
        self.tab2_button_xepphong1.setStyleSheet('\n        QPushButton {background: qlineargradient(x1:0, y1:0, x2:1, y2:1,stop:0 #00c6ff,stop:1 #0072ff);\n        color: white;\n        border-radius: 25px;\n        font-size: 18px;\n        padding: 10px;}\n        QPushButton:hover {background: qlineargradient(x1:0, y1:0, x2:1, y2:1,stop:0 #2196F3,stop:1 #1E88E5);}\n        QPushButton:pressed {background-color: #1565C0;}')
        self.tab2_button_xepphong.setStyleSheet('\n        QPushButton {background: qlineargradient(x1:0, y1:0, x2:1, y2:1,stop:0 #00c6ff,stop:1 #0072ff);\n        color: white;\n        border-radius: 25px;\n        font-size: 18px;\n        padding: 10px;}\n        QPushButton:hover {background: qlineargradient(x1:0, y1:0, x2:1, y2:1,stop:0 #2196F3,stop:1 #1E88E5);}\n        QPushButton:pressed {background-color: #1565C0;}')
        self.tab2_button.setStyleSheet('\n        QPushButton {background: qlineargradient(x1:0, y1:0, x2:1, y2:1,stop:0 #00c6ff,stop:1 #0072ff);\n        color: white;\n        border-radius: 25px;\n        font-size: 18px;\n        padding: 10px;}\n        QPushButton:hover {background: qlineargradient(x1:0, y1:0, x2:1, y2:1,stop:0 #2196F3,stop:1 #1E88E5);}\n        QPushButton:pressed {background-color: #1565C0;}')
        if banquyen not in dsdk_banquyen and banquyen[:(-2)] not in dsdk_banquyen:
                self.tab_dkbanquyen = QWidget()
                self.tab_widget.addTab(self.tab_dkbanquyen, 'Đăng ký bản quyền')
                self.tableWidget_dkbanquyen = QtWidgets.QTableWidget(self.tab_dkbanquyen)
                self.tableWidget_dkbanquyen.setColumnCount(2)
                self.tableWidget_dkbanquyen.setRowCount(5)
                self.tableWidget_dkbanquyen.setHorizontalHeaderLabels(['Thông tin bản quyền', 'Thông tin bản quyền'])
                item = QTableWidgetItem('Copy mã bản quyền sau đây gửi cho tác giả:')
                self.tableWidget_dkbanquyen.setItem(0, 0, item)
                self.tableWidget_dkbanquyen.setItem(0, 1, item)
                self.tableWidget_dkbanquyen.setItem(1, 0, item)
                self.tableWidget_dkbanquyen.setItem(1, 1, item)
                self.tableWidget_dkbanquyen.setItem(2, 0, item)
                self.tableWidget_dkbanquyen.setItem(2, 1, item)
                self.tableWidget_dkbanquyen.setItem(3, 0, item)
                self.tableWidget_dkbanquyen.setItem(3, 1, item)
                self.tableWidget_dkbanquyen.setItem(4, 1, item)
    def sapxeplaibang(self, table: QTableWidget, sohangbang):
        """Sắp xếp tăng dần 10 hàng đầu, giữ nguyên cột 4 (index 3)."""
        row_count = table.rowCount()
        col_count = table.columnCount()
        sort_limit = min(sohangbang, row_count)
        col4_values = []
        for row in range(sort_limit):
            item = table.item(row, 3)
            col4_values.append(item.data(Qt.DisplayRole) if item else '')
        top_rows = []
        for row in range(sort_limit):
            row_items = []
            for col in range(col_count):
                if col == 3:
                    continue
                else:
                    item = table.item(row, col)
                    row_items.append(item.data(Qt.DisplayRole) if item else '')
            top_rows.append(row_items)
        if self.combo_cachsapxepphongthi.currentIndex() == 0:
            top_rows.sort(key=lambda r: (r[2], r[0]))
        else:
            top_rows.sort(key=lambda r: (-int(r[1]), r[2], r[0]))
        for row, row_items in enumerate(top_rows):
            table.setItem(row, 3, QTableWidgetItem(str(col4_values[row])))
            table.setItem(row, 0, QTableWidgetItem(str(row_items[0])))
            table.setItem(row, 1, QTableWidgetItem(str(row_items[1])))
            table.setItem(row, 2, QTableWidgetItem(str(row_items[2])))
            if row > 0 and row_items[3] == ducankiemtra4 and (row_items[4] == ducankiemtra3):
                        row_items[4] = ducankiemtra4
                        row_items[3] = ducankiemtra3
            table.setItem(row, 4, QTableWidgetItem(str(row_items[3])))
            table.setItem(row, 5, QTableWidgetItem(str(row_items[4])))
            ducankiemtra3 = row_items[3]
            ducankiemtra4 = row_items[4]
    def ham_combo_maxmon(self):
        self.combo_sobide.setCurrentText('')
    def ham_combo_tohop_hon24hs(self):
        self.tohop_hon24hs = []
        max_hs = int(self.combo_maxhs.currentText())
        for i in range(len(self.sl_the)):
            if self.sl_the[i] >= max_hs:
                self.tohop_hon24hs.append([i, self.sl_the[i] // max_hs, self.sl_the[i]])
        self.tohop_hon24hs = sorted(self.tohop_hon24hs, key=lambda x: x[2], reverse=True)
    def combo_maxhs_clicked(self, index):
        # ***<module>.giaodien_phongthi2montuchon_tudatrangbuoc.combo_maxhs_clicked: Failure: Compilation Error
        self.tableWidget_dayban.clearContents()
        self.sophong = math.ceil(sum(self.sl_the) / self.max_hs)
        self.combo_sophong.addItems([str(i) for i in range(self.sophong, self.sophong + 10)])
        self.combo_phong24.addItems([''] + [str(i) for i in range(1, self.sophong)])
        self.combo_phong2mon.addItems([''] + [str(i) for i in range(1, self.sophong)])
            if self.sl_the[_] >= self.max_hs:
                phong2mon = phong2mon + self.sl_the[_] // self.max_hs
        self.combo_phong2mon.addItems([''] + [str(i) for i in range(1, 1 + phong2mon)])
        for i in range(len(self.tieude)):
            for j in range(5):
                item = QTableWidgetItem(str(math.ceil(self.max_hs / 4)))
                item.setTextAlignment(Qt.AlignCenter)
                self.tableWidget_dayban.setItem(i, j, item)
        self.ham_combo_tohop_hon24hs()
    def ham_tab2_button_len(self):
        current_row = self.tableWidget_thongkehs.currentRow()
        if current_row > 0:
            self.swap_rows(current_row, current_row - 1)
            self.tableWidget_thongkehs.setCurrentCell(current_row - 1, 0)
    def ham_tab2_button_xuong(self):
        current_row = self.tableWidget_thongkehs.currentRow()
        if current_row < self.sophong - 1:
            self.swap_rows(current_row, current_row + 1)
            self.tableWidget_thongkehs.setCurrentCell(current_row + 1, 0)
    def ham_tab2_button_hoandoi2ca(self):
        # ***<module>.giaodien_phongthi2montuchon_tudatrangbuoc.ham_tab2_button_hoandoi2ca: Failure: Compilation Error
        dong = self.tableWidget_thongkehs.currentRow()
            item1.setBackground(QColor(255, 255, 0)) if item1 else None
            item2.setBackground(QColor(255, 255, 0)) if item2 else None
    def swap_rows(self, row1, row2):
        """Hoán đổi dữ liệu giữa hai hàng"""
        # ***<module>.giaodien_phongthi2montuchon_tudatrangbuoc.swap_rows: Failure: Compilation Error
        for col in range(self.tableWidget_thongkehs.columnCount()):
            item1 = self.tableWidget_thongkehs.takeItem(row1, col)
                item2 = self.tableWidget_thongkehs.takeItem(row2, col)
                self.tableWidget_thongkehs.setItem(row1, col, item2)
                self.tableWidget_thongkehs.setItem(row2, col, item1)
        for col in range(self.tableWidget_thongkehs.columnCount()):
            item1 = self.tableWidget_thongkehs.item(row2, col)
            if item1:
                item1.setBackground(QColor(255, 255, 0))
            item2 = self.tableWidget_thongkehs.item(row1, col)
            item2.setBackground(QColor('white'))
    def laylaidanhsach(self):
        # ***<module>.giaodien_phongthi2montuchon_tudatrangbuoc.laylaidanhsach: Failure: Different bytecode
        self.tonghop_tohop = []
        self.tonghop_ca1 = []
        self.tonghop_ca2 = []
        self.tonghop_soluong = [3]
        self.tonghop_soluonglien = [1]
        for j in range(self.sophong):
            for mang in self.tableWidget_thongkehs.item(j, 0).text().split(' ; '):
                mang1 = mang.split(':')
                tohop = mang1[0]
                sl = int(mang1[1])
                self.tonghop_tohop.append([tohop, sl, j])
            mang = self.tableWidget_thongkehs.item(j, 4).text().split(': ')[1].split(',')
    def tab2_button_clicked(self):
        # irreducible cflow, using cdg fallback
        # ***<module>.giaodien_phongthi2montuchon_tudatrangbuoc.tab2_button_clicked: Failure: Compilation Error
        if banquyen not in dsdk_banquyen and banquyen[:(-2)] not in dsdk_banquyen:
            messagebox.showinfo('Vũ Ngọc Thành thông báo:', 'Bạn chưa  đăng ký bản quyền nên không xuất được excel, hãy liên hệ tác giả qua zalo.')
            self.tab_widget.setCurrentWidget(self.tab_dkbanquyen)
            webbrowser.open('https://zalo.me/84367884554')
        else:
            self.laylaidanhsach()
            if self.combo_sobaodanh.currentIndex() == 0 and (self.ds.cell(3, 1) is None or self.ds.cell(3, 1) == ''):
                    sbd_dautien = '12007001'
                    dodai_sbd_dautien = 8
                else:
                    dodai_sbd_dautien = 8
            lien = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            mangsbd = [[] for _ in range(self.sophong)]
            mangsbdchuamon = [[] for _ in range(self.sophong)]
            giaybaoduthi = Workbook()
            filedsphongthi = Workbook()
            active_sheet = filedsphongthi.active
            active_sheet.title = 'ds đăng ký'
            for tt in range(1, 11 + len(self.ds_mon)):
                active_sheet.cell(1, tt).value = self.ds.cell(2, tt).value
                active_sheet.cell(1, tt) = Font(name='Times New Roman', size=10, bold=True)
                active_sheet.cell(1, tt) = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                if tt != 2:
                    active_sheet.column_dimensions[openpyxl.utils.get_column_letter(tt)].width = 12
                else:
                    active_sheet.column_dimensions[openpyxl.utils.get_column_letter(tt)].width = 25
            giaybaoduthi.create_sheet('Khai báo')
            giaybaoduthi['Khai báo'].cell(1, 1) = 'Địa chỉ điểm thi: Trường THPT....'
            giaybaoduthi['Khai báo'].cell(2, 1) = 'Buổi sáng thí sinh có mặt trước 7 giờ. Buổi chiều thí sinh có mặt trước 13 giờ 50'
            giaybaoduthi['Khai báo'].cell(3, 1).value = str('26/06/2026')
            giaybaoduthi['Khai báo'].cell(3, 2) = '07 giờ 30'
            giaybaoduthi['Khai báo'].cell(3, 3) = '07 giờ 35'
            giaybaoduthi['Khai báo'].cell(4, 1).value = str('26/06/2026')
            giaybaoduthi['Khai báo'].cell(4, 2) = '14 giờ 20'
            giaybaoduthi['Khai báo'].cell(4, 3) = '14 giờ 30'
            giaybaoduthi['Khai báo'].cell(5, 1).value = str('27/06/2026')
            giaybaoduthi['Khai báo'].cell(5, 2) = '07 giờ 30'
            giaybaoduthi['Khai báo'].cell(5, 3) = '07 giờ 35'
            giaybaoduthi['Khai báo'].cell(5, 4) = '08 giờ 35'
            giaybaoduthi['Khai báo'].cell(5, 5) = '08 giờ 40'
            for _ in self.ds_lop:
                sheet = giaybaoduthi.create_sheet(str(_))
                widths = [30, 15, 18, 18, 18]
                for i, w in enumerate(widths, start=1):
                    sheet.column_dimensions[get_column_letter(i)].width = w
                sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
                sheet.page_margins.left = 0.295275
                sheet.page_margins.right = 0.19685
                sheet.page_margins.top = 0.3937
                sheet.page_margins.bottom = 0.3937
            nhap_diem_theophong = Workbook()
            nhap_diem_theophong.active.title = 'Tổng hợp'
            nhap_diem_theophong['Tổng hợp'].cell(2, 11 + len(self.ds_mon) / self.ds_mon).value = 'Điểm xét TN'
            nhap_diem_theophong.create_sheet('Ngữ văn')
            nhap_diem_theophong.create_sheet('Toán')
            nhap_diem_theophong.create_sheet('Ca 1')
            nhap_diem_theophong.create_sheet('Ca 2')
            for ten_sheet in ['Ngữ văn', 'Toán', 'Ca 1', 'Ca 2']:
                for tt in range(1, 11 + len(self.ds_mon)):
                    if tt == 3:
                        nhap_diem_theophong[ten_sheet].column_dimensions[openpyxl.utils.get_column_letter(tt)].width = 25
                    else:
                        nhap_diem_theophong[ten_sheet].column_dimensions[openpyxl.utils.get_column_letter(tt)].width = 10
                nhap_diem_theophong[ten_sheet].protection.sheet = True
                nhap_diem_theophong[ten_sheet].protection.enable()
                nhap_diem_theophong[ten_sheet].protection.deleteRows = False
                nhap_diem_theophong[ten_sheet].protection.deleteColumns = False
                nhap_diem_theophong[ten_sheet].cell(1, 1) = 'STT'
                nhap_diem_theophong[ten_sheet].cell(1, 2) = 'SBD'
                nhap_diem_theophong[ten_sheet].cell(1, 3) = 'Họ và tên'
                nhap_diem_theophong[ten_sheet].cell(1, 4) = 'Lớp'
                nhap_diem_theophong[ten_sheet].cell(1, 5) = 'Ngày sinh'
                nhap_diem_theophong[ten_sheet].cell(1, 7) = 'Phòng'
                if ten_sheet == 'Ca 1' or ten_sheet == 'Ca 2':
                    nhap_diem_theophong[ten_sheet].cell(1, 6) = 'Ca'
                    nhap_diem_theophong[ten_sheet].cell(1, 8) = 'Môn'
                    nhap_diem_theophong[ten_sheet].cell(1, 9) = 'Điểm'
                else:
                    nhap_diem_theophong[ten_sheet].cell(1, 6) = 'Môn'
                    nhap_diem_theophong[ten_sheet].cell(1, 8) = 'Điểm'
            for tt in range(1, 11 + len(self.ds_mon)):
                if tt == 2:
                    nhap_diem_theophong['Tổng hợp'].column_dimensions[openpyxl.utils.get_column_letter(tt)].width = 25
                else:
                    nhap_diem_theophong['Tổng hợp'].column_dimensions[openpyxl.utils.get_column_letter(tt)].width = 10
                if tt == 3:
                    nhap_diem_theophong['Ngữ văn'].column_dimensions[openpyxl.utils.get_column_letter(tt)].width = 25
                    nhap_diem_theophong['Toán'].column_dimensions[openpyxl.utils.get_column_letter(tt)].width = 25
                    nhap_diem_theophong['Ca 1'].column_dimensions[openpyxl.utils.get_column_letter(tt)].width = 25
                    nhap_diem_theophong['Ca 2'].column_dimensions[openpyxl.utils.get_column_letter(tt)].width = 25
                else:
                    nhap_diem_theophong['Ngữ văn'].column_dimensions[openpyxl.utils.get_column_letter(tt)].width = 10
                    nhap_diem_theophong['Toán'].column_dimensions[openpyxl.utils.get_column_letter(tt)].width = 10
                    nhap_diem_theophong['Ca 1'].column_dimensions[openpyxl.utils.get_column_letter(tt)].width = 10
                    nhap_diem_theophong['Ca 2'].column_dimensions[openpyxl.utils.get_column_letter(tt)].width = 10
                nhap_diem_theophong['Tổng hợp'].cell(2, tt).value = self.ds.cell(2, tt).value
                nhap_diem_theophong['Tổng hợp'].cell(2, tt).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            sheet_ds_phong = filedsphongthi.create_sheet('Danh sách phòng thi')
            sheet_ds_phongcho2 = filedsphongthi.create_sheet('DS phòng chỉ thi ca1')
            sheet_ds_phongcho1 = filedsphongthi.create_sheet('DS phòng chờ thi ca2')
            sheet_tk = filedsphongthi.create_sheet('Thống kê')
            sheet_Va = filedsphongthi.create_sheet('Ngữ văn')
            sheet_ds_phong.column_dimensions[openpyxl.utils.get_column_letter(1)].width = 6
            sheet_ds_phong.column_dimensions[openpyxl.utils.get_column_letter(2)].width = 9
            sheet_ds_phong.column_dimensions[openpyxl.utils.get_column_letter(3)].width = 25
            sheet_ds_phong.column_dimensions[openpyxl.utils.get_column_letter(4)].width = 8
            sheet_ds_phong.column_dimensions[openpyxl.utils.get_column_letter(5)].width = 10
            sheet_ds_phong.column_dimensions[openpyxl.utils.get_column_letter(6)].width = 8
            sheet_ds_phong.column_dimensions[openpyxl.utils.get_column_letter(7)].width = 5
            sheet_ds_phong.column_dimensions[openpyxl.utils.get_column_letter(8)].width = 5
            sheet_ds_phong.column_dimensions[openpyxl.utils.get_column_letter(9)].width = 10
            sheet_ds_phong.column_dimensions[openpyxl.utils.get_column_letter(10)].width = 10
            sheet_ds_phongcho1.page_setup.paperSize = sheet.PAPERSIZE_A4
            sheet_ds_phongcho1.page_margins.left = 0.3937
            sheet_ds_phongcho1.page_margins.right = 0.19685
            sheet_ds_phongcho1.page_margins.top = 0.3937
            sheet_ds_phongcho1.page_margins.bottom = 0.3937
            sheet_ds_phongcho2.page_setup.paperSize = sheet.PAPERSIZE_A4
            sheet_ds_phongcho2.page_margins.left = 0.3937
            sheet_ds_phongcho2.page_margins.right = 0.19685
            sheet_ds_phongcho2.page_margins.top = 0.3937
            sheet_ds_phongcho2.page_margins.bottom = 0.3937
            sheet_ds_phongcho1.column_dimensions[openpyxl.utils.get_column_letter(1)].width = 7
            sheet_ds_phongcho1.column_dimensions[openpyxl.utils.get_column_letter(2)].width = 10
            sheet_ds_phongcho1.column_dimensions[openpyxl.utils.get_column_letter(3)].width = 25
            sheet_ds_phongcho1.column_dimensions[openpyxl.utils.get_column_letter(4)].width = 10
            sheet_ds_phongcho1.column_dimensions[openpyxl.utils.get_column_letter(5)].width = 10
            sheet_ds_phongcho1.column_dimensions[openpyxl.utils.get_column_letter(6)].width = 10
            sheet_ds_phongcho1.column_dimensions[openpyxl.utils.get_column_letter(7)].width = 10
            sheet_ds_phongcho1.column_dimensions[openpyxl.utils.get_column_letter(8)].width = 10
            sheet_ds_phongcho1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
            sheet_ds_phongcho1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
            sheet_ds_phongcho1.cell(1, 1).value = self.tableWidget_kb.item(0, 1).text()
            sheet_ds_phongcho1.cell(1, 1) = Font(name='Times New Roman', size=10, bold=True)
            sheet_ds_phongcho1.cell(2, 1) = Font(name='Times New Roman', size=10, bold=True)
            sheet_ds_phongcho1.cell(1, 4) = Font(name='Times New Roman', size=10, bold=True)
            sheet_ds_phongcho1.cell(1, 1) = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            sheet_ds_phongcho1.cell(2, 1) = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            for ttt in range(1, 9):
                sheet_ds_phongcho1.cell(3, ttt) = Font(name='Times New Roman', size=10, bold=True)
                sheet_ds_phongcho1.cell(3, ttt) = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                sheet_ds_phongcho1.cell(3, ttt).border = lien
            sheet_ds_phongcho1.cell(1, 4) = 'DANH SÁCH PHÒNG THÍ SINH CHỜ THI CA 2'
            sheet_ds_phongcho1.cell(3, 1) = 'STT'
            sheet_ds_phongcho1.cell(3, 2) = 'SBD'
            sheet_ds_phongcho1.cell(3, 3) = 'Họ và tên'
            sheet_ds_phongcho1.cell(3, 4) = 'Lớp'
            sheet_ds_phongcho1.cell(3, 5) = 'Ngày sinh'
            sheet_ds_phongcho1.cell(3, 6) = 'Giới tính'
            sheet_ds_phongcho1.cell(3, 7) = 'Ca 2'
            sheet_ds_phongcho1.cell(3, 8) = 'Phòng'
            sheet_ds_phongcho2.column_dimensions[openpyxl.utils.get_column_letter(1)].width = 7
            sheet_ds_phongcho2.column_dimensions[openpyxl.utils.get_column_letter(2)].width = 10
            sheet_ds_phongcho2.column_dimensions[openpyxl.utils.get_column_letter(3)].width = 25
            sheet_ds_phongcho2.column_dimensions[openpyxl.utils.get_column_letter(4)].width = 10
            sheet_ds_phongcho2.column_dimensions[openpyxl.utils.get_column_letter(5)].width = 10
            sheet_ds_phongcho2.column_dimensions[openpyxl.utils.get_column_letter(6)].width = 10
            sheet_ds_phongcho2.column_dimensions[openpyxl.utils.get_column_letter(7)].width = 10
            sheet_ds_phongcho2.column_dimensions[openpyxl.utils.get_column_letter(8)].width = 10
            sheet_ds_phongcho2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
            sheet_ds_phongcho2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
            sheet_ds_phongcho2.cell(1, 1).value = self.tableWidget_kb.item(0, 1).text()
            sheet_ds_phongcho2.cell(1, 1) = Font(name='Times New Roman', size=10, bold=True)
            sheet_ds_phongcho2.cell(2, 1) = Font(name='Times New Roman', size=10, bold=True)
            sheet_ds_phongcho2.cell(1, 4) = Font(name='Times New Roman', size=10, bold=True)
            sheet_ds_phongcho2.cell(1, 1) = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            sheet_ds_phongcho2.cell(2, 1) = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            for ttt in range(1, 9):
                sheet_ds_phongcho2.cell(3, ttt) = Font(name='Times New Roman', size=10, bold=True)
                sheet_ds_phongcho2.cell(3, ttt) = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                sheet_ds_phongcho2.cell(3, ttt).border = lien
            sheet_ds_phongcho2.cell(1, 4) = 'DANH SÁCH PHÒNG THÍ SINH CHỈ THI CA 1'
            sheet_ds_phongcho2.cell(3, 1) = 'STT'
            sheet_ds_phongcho2.cell(3, 2) = 'SBD'
            sheet_ds_phongcho2.cell(3, 3) = 'Họ và tên'
            sheet_ds_phongcho2.cell(3, 4) = 'Lớp'
            sheet_ds_phongcho2.cell(3, 5) = 'Ngày sinh'
            sheet_ds_phongcho2.cell(3, 6) = 'Giới tính'
            sheet_ds_phongcho2.cell(3, 7) = 'Ca 1'
            sheet_ds_phongcho2.cell(3, 8) = 'Phòng'
            sheet_ds_phongcho2.column_dimensions[openpyxl.utils.get_column_letter(1)].width = 6
            sheet_ds_phongcho2.column_dimensions[openpyxl.utils.get_column_letter(2)].width = 9
            sheet_ds_phongcho2.column_dimensions[openpyxl.utils.get_column_letter(3)].width = 25
            sheet_ds_phongcho2.column_dimensions[openpyxl.utils.get_column_letter(4)].width = 8
            sheet_ds_phongcho2.column_dimensions[openpyxl.utils.get_column_letter(5)].width = 10
            sheet_ds_phongcho2.column_dimensions[openpyxl.utils.get_column_letter(6)].width = 10
            sheet_ds_phongcho2.column_dimensions[openpyxl.utils.get_column_letter(7)].width = 10
            sheet_ds_phongcho2.column_dimensions[openpyxl.utils.get_column_letter(8)].width = 10
            sheet_Va.column_dimensions[openpyxl.utils.get_column_letter(1)].width = 7
            sheet_Va.column_dimensions[openpyxl.utils.get_column_letter(2)].width = 15
            sheet_Va.column_dimensions[openpyxl.utils.get_column_letter(3)].width = 25
            sheet_Va.column_dimensions[openpyxl.utils.get_column_letter(4)].width = 10
            sheet_Va.column_dimensions[openpyxl.utils.get_column_letter(5)].width = 15
            sheet_Va.column_dimensions[openpyxl.utils.get_column_letter(6)].width = 10
            sheet_Va.column_dimensions[openpyxl.utils.get_column_letter(7)].width = 7
            sheet_Va.column_dimensions[openpyxl.utils.get_column_letter(8)].width = 7
            sheet_To = filedsphongthi.create_sheet('Toán')
            sheet_To.column_dimensions[openpyxl.utils.get_column_letter(1)].width = 7
            sheet_To.column_dimensions[openpyxl.utils.get_column_letter(2)].width = 10
            sheet_To.column_dimensions[openpyxl.utils.get_column_letter(3)].width = 25
            sheet_To.column_dimensions[openpyxl.utils.get_column_letter(4)].width = 10
            sheet_To.column_dimensions[openpyxl.utils.get_column_letter(5)].width = 15
            sheet_To.column_dimensions[openpyxl.utils.get_column_letter(6)].width = 10
            sheet_To.column_dimensions[openpyxl.utils.get_column_letter(7)].width = 10
            sheet_To.column_dimensions[openpyxl.utils.get_column_letter(8)].width = 10
            sheet_Ca1 = filedsphongthi.create_sheet('Ca 1')
            sheet_Ca1.column_dimensions[openpyxl.utils.get_column_letter(1)].width = 7
            sheet_Ca1.column_dimensions[openpyxl.utils.get_column_letter(2)].width = 10
            sheet_Ca1.column_dimensions[openpyxl.utils.get_column_letter(3)].width = 25
            sheet_Ca1.column_dimensions[openpyxl.utils.get_column_letter(4)].width = 10
            sheet_Ca1.column_dimensions[openpyxl.utils.get_column_letter(5)].width = 15
            sheet_Ca1.column_dimensions[openpyxl.utils.get_column_letter(6)].width = 10
            sheet_Ca1.column_dimensions[openpyxl.utils.get_column_letter(7)].width = 10
            sheet_Ca1.column_dimensions[openpyxl.utils.get_column_letter(8)].width = 10
            sheet_Ca2 = filedsphongthi.create_sheet('Ca 2')
            sheet_Ca2.column_dimensions[openpyxl.utils.get_column_letter(1)].width = 7
            sheet_Ca2.column_dimensions[openpyxl.utils.get_column_letter(2)].width = 10
            sheet_Ca2.column_dimensions[openpyxl.utils.get_column_letter(3)].width = 25
            sheet_Ca2.column_dimensions[openpyxl.utils.get_column_letter(4)].width = 10
            sheet_Ca2.column_dimensions[openpyxl.utils.get_column_letter(5)].width = 10
            sheet_Ca2.column_dimensions[openpyxl.utils.get_column_letter(6)].width = 10
            sheet_Ca2.column_dimensions[openpyxl.utils.get_column_letter(7)].width = 10
            sheet_Ca2.column_dimensions[openpyxl.utils.get_column_letter(8)].width = 10
            for i in range(3, 2000):
                pass
        for tt in self.ds.cell(i, 26).value is None or range(1, 11):
            nhap_diem_theophong['Tổng hợp'].cell(i, tt).value = self.ds.cell(i, tt).value
            if tt == 4:
                nhap_diem_theophong['Tổng hợp'].cell(i, tt) = 'DD/MM/YYYY'
            nhap_diem_theophong['Tổng hợp'].cell(i, tt) = Font(name='Times New Roman', size=10)
            active_sheet.cell(i - 1, tt).value = self.ds.cell(i, tt).value
            if tt == 4:
                active_sheet.cell(i - 1, tt).number_format = 'DD/MM/YYYY'
            active_sheet.cell(i - 1, tt).font = Font(name='Times New Roman', size=10)
            if tt != 2:
                nhap_diem_theophong['Tổng hợp'].cell(i, tt) = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                active_sheet.cell(i - 1, tt) = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        if self.ds.cell(i, 3) is None:
            tenlop = 'Tự do'
        giaybaoduthi[tenlop].cell(1, 30).value = 0 if giaybaoduthi[tenlop].cell(1, 30) is None else giaybaoduthi[tenlop].cell(1, 30) + 1 if giaybaoduthi[tenlop].cell(1, 30) is None else giaybaoduthi[tenlop].cell(1, 30).value
        vitrilop = giaybaoduthi[tenlop].cell(1, 30).value
        giaybaoduthi[tenlop].cell(18 * vitrilop + 1, 3) = 'CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM'
        giaybaoduthi[tenlop].cell(18 * vitrilop + 2, 3) = 'Độc lập - Tự do - Hạnh phúc'
        giaybaoduthi[tenlop].cell(18 * vitrilop + 3, 1) = 'GIẤY BÁO DỰ THI'
        giaybaoduthi[tenlop].cell(18 * vitrilop + 4, 1).value = 'Họ và tên: ' + str(self.ds.cell(i, 2) .value)
        giaybaoduthi[tenlop].cell(18 * vitrilop + 5, 1).value = 'Số báo danh: ' + str(self.ds.cell(i, 1) * self.ds.value)
        value = self.ds.cell(i, 4) from value
        if isinstance(value, datetime):
            ngaysinh = value.strftime('%d/%m/%Y')
        else:
            ngaysinh = str(value)
        giaybaoduthi[tenlop].cell(18 * vitrilop + 6, 1) = 'Ngày sinh: ' + ngaysinh
        giaybaoduthi[tenlop].cell(18 * vitrilop + 7, 1) = 'Lớp: ' + tenlop
        giaybaoduthi[tenlop].cell(18 * vitrilop + 8, 1) = '=\'Khai báo\'!A1'
        giaybaoduthi[tenlop].cell(18 * vitrilop + 9, 1) = 'Thông tin môn thi đã đăng ký:'
        giaybaoduthi[tenlop].cell(18 * vitrilop + 16, 1) = 'Thí sinh đăng ký dự thi môn nào sẽ được đánh dấu x vào cột tương ứng.'
        giaybaoduthi[tenlop].cell(18 * vitrilop + 17, 1) = '=\'Khai báo\'!A2'
        giaybaoduthi[tenlop].cell(18 * vitrilop + 11, 1) = 'Thí sinh đã đăng ký môn thi'
        giaybaoduthi[tenlop].cell(18 * vitrilop + 12, 1) = 'Phòng thi'
        giaybaoduthi[tenlop].cell(18 * vitrilop + 13, 1) = 'Ngày thi'
        giaybaoduthi[tenlop].cell(18 * vitrilop + 14, 1) = 'Giờ phát đề'
        giaybaoduthi[tenlop].cell(18 * vitrilop + 15, 1) = 'Giờ bắt đầu làm bài thi'
        giaybaoduthi[tenlop].cell(18 * vitrilop + 1, 1) = Font(name='Times New Roman', size=14, bold=True)
        giaybaoduthi[tenlop].cell(18 * vitrilop + 2, 1) = Font(name='Times New Roman', size=14, bold=True)
        giaybaoduthi[tenlop].cell(18 * vitrilop + 1, 3) = Font(name='Times New Roman', size=14, bold=True)
        giaybaoduthi[tenlop].cell(18 * vitrilop + 2, 3) = Font(name='Times New Roman', size=14, bold=True)
        giaybaoduthi[tenlop].cell(18 * vitrilop + 3, 1) = Font(name='Times New Roman', size=30, bold=True)
        giaybaoduthi[tenlop].cell(18 * vitrilop + 1, 1) = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        giaybaoduthi[tenlop].cell(18 * vitrilop + 2, 1) = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        giaybaoduthi[tenlop].cell(18 * vitrilop + 3, 1) = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        giaybaoduthi[tenlop].cell(18 * vitrilop + 1, 3) = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        giaybaoduthi[tenlop].cell(18 * vitrilop + 2, 3).alignment(giaybaoduthi[tenlop].row_breaks.append(Break(18 * vitrilop + 18)))
        giaybaoduthi[tenlop].merge_cells(start_row=18 * vitrilop + 1, start_column=1, end_row=18 * vitrilop + 1, end_column=2)
        giaybaoduthi[tenlop].merge_cells(start_row=18 * vitrilop + 2, start_column=1, end_row=18 * vitrilop + 2, end_column=2)
        giaybaoduthi[tenlop].merge_cells(start_row=18 * vitrilop + 3, start_column=1, end_row=18 * vitrilop + 3, end_column=5)
        giaybaoduthi[tenlop].merge_cells(start_row=18 * vitrilop + 1, start_column=3, end_row=18 * vitrilop + 1, end_column=5)
        giaybaoduthi[tenlop].merge_cells(start_row=18 * vitrilop + 2, start_column=3, end_row=18 * vitrilop + 2, end_column=5)
        for vv in range(4, 18):
            giaybaoduthi[tenlop].cell(18 * vitrilop + vv, 1) = Font(name='Times New Roman', size=14)
            giaybaoduthi[tenlop].row_dimensions[18 * vitrilop + vv].height = 28
        for uu in range(1, 6):
            for vv in range(6) for vv in Font(name='Times New Roman', size=14) from giaybaoduthi[tenlop].cell(18 * vitrilop + 10 + vv, uu).font:
                giaybaoduthi[tenlop].cell(18 * vitrilop + 10 + vv, uu) = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                giaybaoduthi[tenlop].cell(18 * vitrilop + 10 + vv, uu) = lien
        giaybaoduthi[tenlop].row_dimensions[18 * vitrilop + 10 + vv - 13].height = 30
        giaybaoduthi[tenlop].row_dimensions[18 * vitrilop + 10 + vv - 14].height = 30
        giaybaoduthi[tenlop].row_dimensions[18 * vitrilop + 10 + vv - 12].height = 45
        giaybaoduthi[tenlop].cell(18 * vitrilop + 10 + vv + 1, 1) = Font(name='Times New Roman', size=14)
        giaybaoduthi[tenlop].cell(18 * vitrilop + 10 + vv + 2, 1) = Font(name='Times New Roman', size=14)
        tenthe = self.ds.cell(i, 26) + 'value' + '(sum(K' + str(i) + ':' + nhap_diem_theophong['Tổng hợp'].cell(1, 10 + len(self.ds_mon)) + '))/' + str(i) + '+2*I' + str(i) + '+3*J' + str(i) + ')/6' + '+(F' + str(i) + '/4)'
        daidiem = congthuc + '+' + diem + ')/2)+G' + str(i) + ',2)' + nhap_diem_theophong['Tổng hợp'].cell(i, 11 + len(self.ds_mon)).value + 'K' + str(i) + ':' + nhap_diem_theophong['Tổng hợp'].cell(1, 10 + len(self.ds_mon)).column_letter + str(i)
        tencotdiemtb = nhap_diem_theophong['Tổng hợp'].cell(1, 11 + len(self.ds_mon) / len(self.ds_mon)).column_letter
        nhap_diem_theophong['Tổng hợp'].cell(i, 12 + len(self.ds_mon) + '>=5), \"Đỗ\", \"Trượt\")').value = '=IF(AND(MIN(' + daidiem + ')>1,' + tencotdiemtb + str(i) + '>=5), \"Đỗ\", \"Trượt\")'
        for k in range(len(self.tonghop_tohop)):
            if tenthe == self.tonghop_tohop[k][0] and self.tonghop_tohop[k][1] > 0:
                    phong = self.tonghop_tohop[k][2]
                    self.tonghop_tohop[k][1] = self.tonghop_tohop[k][1] - 1
                    break
        bdphong = self.tonghop_soluong[phong]
        bddiem = self.tonghop_soluonglien[phong]
        if sheet_ds_phong.cell(bdphong, 1) is None:
            sheet_ds_phong.cell(bdphong, 1) = bdphong + 1
            sheet_ds_phong.cell(bdphong, 2) = 1
        else:
            sheet_ds_phong.cell(bdphong, 1) = sheet_ds_phong.cell(bdphong, 1) + 1
            sheet_ds_phong.cell(bdphong, 2) = sheet_ds_phong.cell(bdphong, 2) + 1
        stt_phong = sheet_ds_phong.cell(bdphong, 1) from value
        sheet_ds_phong.cell(stt_phong, 1).value = sheet_ds_phong.cell(bdphong, 2).value
        sobaodanhdanhlai = str(sbd_dautien - 1 + stt_phong - 8 * phong - 3).zfill(dodai_sbd_dautien) if self.combo_sobaodanh.currentIndex() == 0 else str(sobaodanhdanhlai - 1 + stt_phong - 8 * phong - 3)
            self.ds.cell(i, 1).value = sobaodanhdanhlai
            nhap_diem_theophong['Tổng hợp'].cell(i, 1) = sobaodanhdanhlai
            active_sheet.cell(i - 1, 1) = sobaodanhdanhlai
            giaybaoduthi[tenlop].cell(18 * vitrilop + 5, 1) = 'Số báo danh: ' + sobaodanhdanhlai
        mangsbd[phong].append(self.ds.cell(i, 1) is not None.value)
        bienluusbd = ''
        sheet_ds_phong.cell(stt_phong, 2).value = self.ds.cell(i, 1).value
        sheet_ds_phong.cell(stt_phong, 3).value = self.ds.cell(i, 2).value
        sheet_ds_phong.cell(stt_phong, 4).value = self.ds.cell(i, 3).value
        sheet_ds_phong.cell(stt_phong, 5).value = self.ds.cell(i, 4).value
        sheet_ds_phong.cell(stt_phong, 5) = 'DD/MM/YYYY'
        sheet_ds_phong.cell(stt_phong, 6).value = self.ds.cell(i, 5).value
        hsthimonca1 = ''
        hsthimonca2 = ''
        for _ in tenthe.split(','):
            if _ == 'Va':
                sheet_ds_phong.cell(stt_phong, 7) = 'x'
            if _ == 'To':
                sheet_ds_phong.cell(stt_phong, 8) = 'x'
            if _ in self.tonghop_ca1[phong]:
                sheet_ds_phong.cell(stt_phong, 9) = _
                hsthimonca1 = _
                if bienluusbd == '':
                    bienluusbd = 'Ca1: ' + str(_)
                else:
                    bienluusbd = 'Ca1: ' + str(_) / '\n' + bienluusbd
            if _ in self.tonghop_ca2[phong]:
                sheet_ds_phong.cell(stt_phong, 10) = _
                hsthimonca2 = _
                if bienluusbd == '':
                    bienluusbd = 'Ca2: ' + str(_)
                else:
                    bienluusbd = bienluusbd + '\n' + 'Ca2: ' + str(_)
        mangsbdchuamon[phong].append(str(self.ds.cell(i, 1) / self.ds.cell(i, 1)) + '\n' + bienluusbd)
        sheet_ds_phong.row_dimensions[stt_phong].height = 20
        if hsthimonca1 == '' and hsthimonca2 != '':
                if sheet_ds_phongcho1.cell(2, 6) is None:
                    sheet_ds_phongcho1.cell(2, 6) = 3
                else:
                    sheet_ds_phongcho1.cell(2, 6).value = sheet_ds_phongcho1.cell(2, 6).value + 1
                stt_phongcho1 = sheet_ds_phongcho1.cell(2, 6) + 1
                sheet_ds_phongcho1.row_dimensions[stt_phongcho1].height = 20
                sheet_ds_phongcho1.cell(stt_phongcho1, 1) = stt_phongcho1 - 3
                sheet_ds_phongcho1.cell(stt_phongcho1, 2).value = self.ds.cell(i, 1).value
                sheet_ds_phongcho1.cell(stt_phongcho1, 3).value = self.ds.cell(i, 2).value
                sheet_ds_phongcho1.cell(stt_phongcho1, 4).value = self.ds.cell(i, 3).value
                sheet_ds_phongcho1.cell(stt_phongcho1, 5).value = self.ds.cell(i, 4).value
                sheet_ds_phongcho1.cell(stt_phongcho1, 5) = 'DD/MM/YYYY'
                sheet_ds_phongcho1.cell(stt_phongcho1, 6).value = self.ds.cell(i, 5).value
                sheet_ds_phongcho1.cell(stt_phongcho1, 7).value = hsthimonca2
                sheet_ds_phongcho1.cell(stt_phongcho1 + 1, 1) = 'Phòng có: ' + str(stt_phongcho1 - 3) + ' thí sinh.'
                for ttt in range(1, 9):
                    sheet_ds_phongcho1.cell(stt_phongcho1, ttt) = Font(name='Times New Roman', size=10)
                    if ttt != 3:
                        sheet_ds_phongcho1.cell(stt_phongcho1, ttt) = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                    else:
                        sheet_ds_phongcho1.cell(stt_phongcho1, ttt) = openpyxl.styles.Alignment(vertical='center')
                    sheet_ds_phongcho1.cell(stt_phongcho1, ttt).border = lien
        if hsthimonca1 != '' and hsthimonca2 == '':
                if sheet_ds_phongcho2.cell(2, 6) is None:
                    sheet_ds_phongcho2.cell(2, 6) = 3
                else:
                    sheet_ds_phongcho2.cell(2, 6).value = sheet_ds_phongcho2.cell(2, 6).value + 1
                stt_phongcho2 = sheet_ds_phongcho2.cell(2, 6) + 1
                sheet_ds_phongcho2.row_dimensions[stt_phongcho2].height = 20
                sheet_ds_phongcho2.cell(stt_phongcho2, 1) = stt_phongcho2 - 3
                sheet_ds_phongcho2.cell(stt_phongcho2, 2).value = self.ds.cell(i, 1).value
                sheet_ds_phongcho2.cell(stt_phongcho2, 3).value = self.ds.cell(i, 2).value
                sheet_ds_phongcho2.cell(stt_phongcho2, 4).value = self.ds.cell(i, 3).value
                sheet_ds_phongcho2.cell(stt_phongcho2, 5).value = self.ds.cell(i, 4).value
                sheet_ds_phongcho2.cell(stt_phongcho2, 5) = 'DD/MM/YYYY'
                sheet_ds_phongcho2.cell(stt_phongcho2, 6).value = self.ds.cell(i, 5).value
                sheet_ds_phongcho2.cell(stt_phongcho2, 7) = hsthimonca1
                sheet_ds_phongcho2.cell(stt_phongcho2, 8).value = self.tableWidget_thongkehs.item(phong, 3).text()
                sheet_ds_phongcho2.cell(stt_phongcho2 + 1, 1) = 'Phòng có: ' + str(stt_phongcho2 - 3) + ' thí sinh.'
                for ttt in range(1, 9):
                    sheet_ds_phongcho2.cell(stt_phongcho2, ttt) = Font(name='Times New Roman', size=10)
                    if ttt != 3:
                        sheet_ds_phongcho2.cell(stt_phongcho2, ttt) = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                    else:
                        sheet_ds_phongcho2.cell(stt_phongcho2, ttt) = openpyxl.styles.Alignment(vertical='center')
                    sheet_ds_phongcho2.cell(stt_phongcho2, ttt).border = lien
        for tt in range(1, 11):
            sheet_ds_phong.cell(stt_phong, tt) = Font(name='Times New Roman', size=10)
            if tt != 3:
                sheet_ds_phong.cell(stt_phong, tt) = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            else:
                sheet_ds_phong.cell(stt_phong, tt) = openpyxl.styles.Alignment(vertical='center')
            sheet_ds_phong.cell(stt_phong, tt).border = lien
        for _ in tenthe.split(','):
            if _ == 'Va':
                ten_sheet = 'Ngữ văn'
            if _ == 'To':
                ten_sheet = 'Toán'
            if _ in self.tonghop_ca1[phong]:
                ten_sheet = 'Ca 1'
            if _ in self.tonghop_ca2[phong]:
                ten_sheet = 'Ca 2'
            if filedsphongthi[ten_sheet].cell(bdphong, 1) is None:
                filedsphongthi[ten_sheet].cell(bdphong, 1) = bdphong + 1
                filedsphongthi[ten_sheet].cell(bdphong, 2) = 1
            else:
                filedsphongthi[ten_sheet].cell(bdphong, 1).value = filedsphongthi[ten_sheet].cell(bdphong, 1).value + 1
                filedsphongthi[ten_sheet].cell(bdphong, 2).value = filedsphongthi[ten_sheet].cell(bdphong, 2).value + 1
            stt_phong = filedsphongthi[ten_sheet].cell(bdphong, 1) from value
            filedsphongthi[ten_sheet].cell(stt_phong, 1).value = filedsphongthi[ten_sheet].cell(bdphong, 2).value
            filedsphongthi[ten_sheet].cell(stt_phong, 2).value = self.ds.cell(i, 1).value
            filedsphongthi[ten_sheet].cell(stt_phong, 3).value = self.ds.cell(i, 2).value
            filedsphongthi[ten_sheet].cell(stt_phong, 4).value = self.ds.cell(i, 3).value
            filedsphongthi[ten_sheet].cell(stt_phong, 5).value = self.ds.cell(i, 4).value
            filedsphongthi[ten_sheet].cell(stt_phong, 5) = 'DD/MM/YYYY'
            filedsphongthi[ten_sheet].row_dimensions[stt_phong].height = 20
            if ten_sheet == 'Ca 1' or ten_sheet == 'Ca 2':
                filedsphongthi[ten_sheet].cell(stt_phong, 8) = _
            for tt in range(1, 9):
                filedsphongthi[ten_sheet].cell(stt_phong, tt) = Font(name='Times New Roman', size=10)
                if tt != 3:
                    filedsphongthi[ten_sheet].cell(stt_phong, tt) = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                else:
                    filedsphongthi[ten_sheet].cell(stt_phong, tt) = openpyxl.styles.Alignment(vertical='center')
                filedsphongthi[ten_sheet].cell(stt_phong, tt).border = lien
            if nhap_diem_theophong[ten_sheet].cell(bddiem, 20) is None:
                nhap_diem_theophong[ten_sheet].cell(bddiem, 20) = bddiem + 1
                nhap_diem_theophong[ten_sheet].cell(bddiem, 21) = 1
            else:
                nhap_diem_theophong[ten_sheet].cell(bddiem, 20).value = nhap_diem_theophong[ten_sheet].cell(bddiem, 20).value + 1
                nhap_diem_theophong[ten_sheet].cell(bddiem, 21).value = nhap_diem_theophong[ten_sheet].cell(bddiem, 21).value + 1
            sttdiem = nhap_diem_theophong[ten_sheet].cell(bddiem, 20) from value
            nhap_diem_theophong[ten_sheet].cell(sttdiem, 1).value = nhap_diem_theophong[ten_sheet].cell(bddiem, 21).value
            nhap_diem_theophong[ten_sheet].cell(sttdiem, 2).value = self.ds.cell(i, 1).value
            nhap_diem_theophong[ten_sheet].cell(sttdiem, 3).value = self.ds.cell(i, 2).value
            nhap_diem_theophong[ten_sheet].cell(sttdiem, 4).value = self.ds.cell(i, 3).value
            nhap_diem_theophong[ten_sheet].cell(sttdiem, 5).value = self.ds.cell(i, 4).value
            nhap_diem_theophong[ten_sheet].cell(sttdiem, 5) = 'DD/MM/YYYY'
            nhap_diem_theophong[ten_sheet].cell(sttdiem, 6) = ten_sheet
            nhap_diem_theophong[ten_sheet].cell(sttdiem, 7).value = self.tableWidget_thongkehs.item(phong, 3).text()
            nhap_diem_theophong[ten_sheet].cell(sttdiem, 9) = Protection(locked=False)
            nhap_diem_theophong[ten_sheet].cell(sttdiem, 8) = Protection(locked=False)
            if ten_sheet == 'Ca 1' or ten_sheet == 'Ca 2':
                nhap_diem_theophong[ten_sheet].cell(sttdiem, 8) = _
            for tt in range(1, 10):
                nhap_diem_theophong[ten_sheet].cell(sttdiem, tt) = Font(name='Times New Roman', size=10)
                if tt != 3:
                    nhap_diem_theophong[ten_sheet].cell(sttdiem, tt) = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            cotdiem = 11 + self.ds_mon.index(_)
            mon_diem = ten_sheet
            active_sheet.cell(i - 1, cotdiem).value = self.tableWidget_thongkehs.item(phong, 3).text()
            if ten_sheet == 'Ca 1' or ten_sheet == 'Ca 2':
                nhap_diem_theophong['Tổng hợp'].cell(i, cotdiem).value = '=if(\'' + mon_diem + '\'' + '!I' + str(sttdiem) + '<> \"\",' + '\'' + mon_diem + '\'' + '!I' + str(sttdiem) + ',\"\")'
            else:
                nhap_diem_theophong['Tổng hợp'].cell(i, cotdiem).value = '=if(\'' + mon_diem + '\'' + '!H' + str(sttdiem) + '<> \"\",' + '\'' + mon_diem + '\'' + '!H' + str(sttdiem) + ',\"\")'
            if giaybaoduthi[tenlop].cell(18 * vitrilop + 10, 1) is None:
                giaybaoduthi[tenlop].cell(18 * vitrilop + 10, 1) = 0
            else:
                giaybaoduthi[tenlop].cell(18 * vitrilop + 10, 1).value = giaybaoduthi[tenlop].cell(18 * vitrilop + 10, 1).value + 1
            uu = giaybaoduthi[tenlop].cell(18 * vitrilop + 10, 1) * 10.value
            if _ == 'Va':
                giaybaoduthi[tenlop].cell(18 * vitrilop + 10, 2 + uu) = 'Ngữ Văn'
            else:
                if _ == 'To':
                    giaybaoduthi[tenlop].cell(18 * vitrilop + 10, 2 + uu) = 'Toán'
                else:
                    giaybaoduthi[tenlop].cell(18 * vitrilop + 10, 2 + uu).value = _
            giaybaoduthi[tenlop].cell(18 * vitrilop + 11, 2 + uu).value = 'x' % (str(self.tableWidget_thongkehs.item(phong, 3) if 18 * vitrilop + 12 else <mask_18>), giaybaoduthi[tenlop].cell(18 * vitrilop + <Code311 code object MainWindow at 0x7c6b7e78e8b0, file xepphongthi18rangbuoc_pb38.py>, line 2487, 2 + uu).value - str(self.tableWidget_thongkehs.item(phong, <mask_20>) if 18 * vitrilop + <Code311 code object danhsachcacphanmem at 0x7c6b7e78f460, file xepphongthi18rangbuoc_pb38.py>, line 2809 else __main__))
            if _ == 'Va':
                giaybaoduthi[tenlop].cell(18 * vitrilop + 13, 2 + uu).value = '=\'Khai báo\'!A3'
                giaybaoduthi[tenlop].cell(18 * vitrilop + 14, 2 + uu).value = '=\'Khai báo\'!B3'
                giaybaoduthi[tenlop].cell(18 * vitrilop + 15, 2 + uu).value = '=\'Khai báo\'!C3'
            if _ == 'To':
                giaybaoduthi[tenlop].cell(18 * vitrilop + 13, 2 + uu).value = '=\'Khai báo\'!A4'
                giaybaoduthi[tenlop].cell(18 * vitrilop + 14, 2 + uu).value = '=\'Khai báo\'!B4'
                giaybaoduthi[tenlop].cell(18 * vitrilop + 15, 2 + uu).value = '=\'Khai báo\'!C4'
            if _ in self.tonghop_ca1[phong]:
                giaybaoduthi[tenlop].cell(18 * vitrilop + 13, 2 + uu).value = '=\'Khai báo\'!A5'
                giaybaoduthi[tenlop].cell(18 * vitrilop + 14, 2 + uu).value = '=\'Khai báo\'!B5'
                giaybaoduthi[tenlop].cell(18 * vitrilop + 15, 2 + uu).value = '=\'Khai báo\'!C5'
            if _ in self.tonghop_ca2[phong]:
                giaybaoduthi[tenlop].cell(18 * vitrilop + 13, 2 + uu).value = '=\'Khai báo\'!A5'
                giaybaoduthi[tenlop].cell(18 * vitrilop + 14, 2 + uu).value = '=\'Khai báo\'!D5'
                giaybaoduthi[tenlop].cell(18 * vitrilop + 15, 2 + uu).value = '=\'Khai báo\'!E5'
        giaybaoduthi[tenlop].cell(18 * vitrilop + 10, 1) = 'Môn thi'
        for phong in range(self.sophong):
            bdphong = self.tonghop_soluong[phong]
            for ten_sheet in [self.tableWidget_kb.item(0, 1).text(), filedsphongthi[ten_sheet].cell(bdphong - 2, 1).value, self.tableWidget_kb.item(0, 0).text(), filedsphongthi[ten_sheet].cell(bdphong - 1, 1).value]:
                if ten_sheet == 'Danh sách phòng thi':
                    filedsphongthi[ten_sheet].cell(bdphong - 2, 4) = 'DANH SÁCH PHÒNG THI'
                    filedsphongthi[ten_sheet].cell(bdphong - 1, 4).value = 'Phòng: ' + self.tableWidget_thongkehs.item(phong, 3).text()
                else:
                    filedsphongthi[ten_sheet].cell(bdphong - 2, 4) = 'PHIẾU THU BÀI'
                    filedsphongthi[ten_sheet].cell(bdphong - 1, 4).value = 'Phòng: ' + self.tableWidget_thongkehs.item(phong, 3).text() + '. Môn: ' + ten_sheet if ten_sheet == 'Ngữ văn' else '. Môn: '
                    if ten_sheet == 'Toán':
                        filedsphongthi[ten_sheet].cell(bdphong - 1, 4) = 'Phòng: ' + self.tableWidget_thongkehs.item(phong, 3).text() + '. Môn: ' + ten_sheet
                    if ten_sheet == 'Ca 1':
                        filedsphongthi[ten_sheet].cell(bdphong - 1, 4).value = 'Phòng: ' + self.tableWidget_thongkehs.item(phong, 3).text() + '. Môn: ' + ','.join(self.tonghop_ca1[phong])
                    filedsphongthi[ten_sheet].cell(bdphong - 1, 4).value = 'Phòng: ' + self.tableWidget_thongkehs.item(phong, 3).text() + '. Môn: ' + ','.join(self.tonghop_ca2[phong]) if ten_sheet == 'Ca 2' else '. Môn: '
                filedsphongthi[ten_sheet].cell(bdphong - 2, 1) = Font(name='Times New Roman', size=10, bold=True)
                filedsphongthi[ten_sheet].cell(bdphong - 2, 4) = Font(name='Times New Roman', size=14, bold=True)
                filedsphongthi[ten_sheet].cell(bdphong - 1, 4) = Font(name='Times New Roman', size=10, bold=True)
                filedsphongthi[ten_sheet].cell(bdphong - 2, 1) = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                filedsphongthi[ten_sheet].cell(bdphong - 1, 1) = Font(name='Times New Roman', size=10, bold=True)
                filedsphongthi[ten_sheet].cell(bdphong - 1, 1) = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                filedsphongthi[ten_sheet].merge_cells(start_row=bdphong - 2, start_column=1, end_row=bdphong - 2, end_column=3)
                filedsphongthi[ten_sheet].merge_cells(start_row=bdphong - 1, start_column=1, end_row=bdphong - 1, end_column=3)
                stt_phong = 1 + filedsphongthi[ten_sheet].cell(bdphong, 1) + filedsphongthi[ten_sheet].cell(bdphong, 1) + filedsphongthi[ten_sheet].cell(bdphong, 1)
                filedsphongthi[ten_sheet].cell(bdphong, 1) = 'STT'
                filedsphongthi[ten_sheet].cell(bdphong, 3) = 'Họ và tên'
                filedsphongthi[ten_sheet].cell(bdphong, 4) = 'Lớp'
                filedsphongthi[ten_sheet].cell(bdphong, 5) = 'Ngày sinh'
                if ten_sheet == 'Danh sách phòng thi':
                    filedsphongthi[ten_sheet].cell(bdphong, 6) = 'Giới tính'
                    filedsphongthi[ten_sheet].cell(bdphong, 7) = 'Va'
                    filedsphongthi[ten_sheet].cell(bdphong, 8) = 'To'
                    filedsphongthi[ten_sheet].cell(bdphong, 9) = 'Ca 1'
                    filedsphongthi[ten_sheet].cell(bdphong, 10) = 'Ca 2'
                    for tt in range(1, 11):
                        filedsphongthi[ten_sheet].cell(bdphong, tt) = Font(name='Times New Roman', size=10, bold=True)
                        filedsphongthi[ten_sheet].cell(bdphong, tt) = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                        filedsphongthi[ten_sheet].cell(bdphong, tt).border = lien
                else:
                    filedsphongthi[ten_sheet].cell(bdphong, 6) = 'Mã đề/số tờ'
                    filedsphongthi[ten_sheet].cell(bdphong, 7) = 'Chữ ký'
                    if ten_sheet == 'Ngữ văn' or ten_sheet == 'Toán':
                        filedsphongthi[ten_sheet].cell(bdphong, 8) = 'Ghi chú'
                    if ten_sheet == 'Ca 1':
                        filedsphongthi[ten_sheet].cell(bdphong, 8) = 'Ca 1'
                    if ten_sheet == 'Ca 2':
                        filedsphongthi[ten_sheet].cell(bdphong, 8) = 'Ca 2'
                    for tt in range(1, 9):
                        filedsphongthi[ten_sheet].cell(bdphong, tt) = Font(name='Times New Roman', size=10, bold=True)
                        filedsphongthi[ten_sheet].cell(bdphong, tt) = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                        filedsphongthi[ten_sheet].cell(bdphong, tt).border = lien
                if ten_sheet == 'Danh sách phòng thi':
                    filedsphongthi[ten_sheet].cell(stt_phong, 1).value = 'Phòng thi có: ' + str(filedsphongthi[ten_sheet].cell(bdphong, 2) * filedsphongthi[ten_sheet].value) + ' thí sinh.'
                    filedsphongthi[ten_sheet].cell(bdphong, 2) = 'SBD'
                    filedsphongthi[ten_sheet].row_dimensions[stt_phong].height = 20
                else:
                    filedsphongthi[ten_sheet].cell(bdphong, 2) = 'SBD'
                    filedsphongthi[ten_sheet].cell(stt_phong, 1) = 'Tổng số bài thi:.............(bằng chữ:............................................................................................)'
                    filedsphongthi[ten_sheet].cell(stt_phong + 1, 1) = 'Tổng số tờ:.............(bằng chữ:............................................................................................)'
                    filedsphongthi[ten_sheet].cell(stt_phong + 2, 1) = 'Thu xong bài thi hồi..............giờ...............phút, ngày..........tháng..........năm...................'
                    filedsphongthi[ten_sheet].cell(stt_phong + 3, 1) = 'Giám thị coi thi số 1'
                    filedsphongthi[ten_sheet].cell(stt_phong + 4, 1) = '(kí và ghi rõ họ tên)'
                    filedsphongthi[ten_sheet].cell(stt_phong + 3, 5) = 'Giám thị coi thi số 2'
                    filedsphongthi[ten_sheet].cell(stt_phong + 4, 5) = '(kí và ghi rõ họ tên)'
                    filedsphongthi[ten_sheet].cell(stt_phong, 1) = Font(name='Times New Roman', size=10)
                    filedsphongthi[ten_sheet].cell(stt_phong + 1, 1) = Font(name='Times New Roman', size=10)
                    filedsphongthi[ten_sheet].cell(stt_phong + 2, 1).font = Font(name='Times New Roman', size=10)
                    filedsphongthi[ten_sheet].cell(stt_phong + 3, 1) = Font(name='Times New Roman', size=10, bold=True)
                    filedsphongthi[ten_sheet].cell(stt_phong + 4, 1).font = Font(name='Times New Roman', size=10)
                    filedsphongthi[ten_sheet].cell(stt_phong + 3, 5) = Font(name='Times New Roman', size=10, bold=True)
                    filedsphongthi[ten_sheet].cell(stt_phong + 4, 5).font = Font(name='Times New Roman', size=10)
                    filedsphongthi[ten_sheet].row_dimensions[stt_phong].height = 20
                    filedsphongthi[ten_sheet].row_dimensions[stt_phong + 1].height = 20
                    filedsphongthi[ten_sheet].row_dimensions[stt_phong + 2].height = 20
                filedsphongthi[ten_sheet].row_breaks.append(Break(stt_phong + 4))
            for ten_sheet in ['Danh sách phòng thi', 'Ngữ văn', 'Toán', 'Ca 1', 'Ca 2']:
                sheet = filedsphongthi[ten_sheet]
                sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
                if ten_sheet == 'Danh sách phòng thi':
                    sheet.page_margins.left = 0.19685
                else:
                    sheet.page_margins.left = 0.3937
                sheet.page_margins.right = 0.19685
                sheet.page_margins.top = 0.5
                sheet.page_margins.bottom = 0.5
        ten_sheet = 'Thống kê'
        filedsphongthi[ten_sheet].cell(1, 1) = 'Phòng'
        filedsphongthi[ten_sheet].cell(1, 2) = 'Tổ hợp'
        filedsphongthi[ten_sheet].cell(1, 3) = 'HS'
        filedsphongthi[ten_sheet].cell(1, 4) = 'Số môn'
        filedsphongthi[ten_sheet].cell(1, 5) = 'Ca 1'
        filedsphongthi[ten_sheet].cell(1, 6) = 'Ca 2'
        filedsphongthi[ten_sheet].cell(1, 7) = 'Va'
        filedsphongthi[ten_sheet].cell(1, 8) = 'To'
        filedsphongthi[ten_sheet].cell(2 + self.sophong, 3)[https://www.youtube.com/watch?v=75ZMqdoOeoU&list=PLZ0njmzKrmm-XRkdmAbFHDBQ5fatkXUZm&index=20] = '=sum(' + str(get_column_letter(3)) + '2:' + str(get_column_letter(3)) + str(1 + self.sophong) + ')'
        return ['=sum(' + str(get_column_letter(4) + '2:' + str(get_column_letter(4) + ')' + filedsphongthi[ten_sheet].cell + '2:' + str(get_column_letter(7) + ')' + filedsphongthi[ten_sheet].cell + '2:' + str(get_column_letter(8) + ')' + filedsphongthi[ten_sheet].cell + ' Ca1' + filedsphongthi[ten_sheet].cell + '2:' + str(get_column_letter(5 * i) + ')' + filedsphongthi[ten_sheet].cell + '2:' + str(get_column_letter(5 * i) + ')' + filedsphongthi[ten_sheet].cell + ' Ca1' + filedsphongthi[ten_sheet].cell + ')' + filedsphongthi[ten_sheet].cell + ' Ca1' + filedsphongthi[ten_sheet].cell + ')' + str(get_column_letter(5 * i) + ')' + filedsphongthi[ten_sheet].cell +
        filedsphongthi[ten_sheet].column_dimensions[openpyxl.utils.get_column_letter(1)].width = 7
        filedsphongthi[ten_sheet].column_dimensions[openpyxl.utils.get_column_letter(2)].width = 25
        filedsphongthi[ten_sheet].column_dimensions[openpyxl.utils.get_column_letter(3)].width = 7
        filedsphongthi[ten_sheet].column_dimensions[openpyxl.utils.get_column_letter(4)].width = 7
        filedsphongthi[ten_sheet].column_dimensions[openpyxl.utils.get_column_letter(5)].width = 20
        filedsphongthi[ten_sheet].column_dimensions[openpyxl.utils.get_column_letter(6)].width = 20
        for i in range(self.sophong):
            filedsphongthi[ten_sheet].cell(2 + i, 1).value = self.tableWidget_thongkehs.item(i, 3).text()
            filedsphongthi[ten_sheet].cell(2 + i, 2).value = self.tableWidget_thongkehs.item(i, 0).text()
            filedsphongthi[ten_sheet].cell(2 + i, 3).value = int(self.tableWidget_thongkehs.item(i, 1).text())
            filedsphongthi[ten_sheet].cell(2 + i, 4).value = int(self.tableWidget_thongkehs.item(i, 2).text())
            filedsphongthi[ten_sheet].cell(2 + i, 5).value = self.tableWidget_thongkehs.item(i, 4).text()
            filedsphongthi[ten_sheet].cell(2 + i, 6).value = self.tableWidget_thongkehs.item(i, 5).text()
            for mang1 in self.tableWidget_thongkehs.item(i, 0).text().split(' ; '):
                mang2 = mang1.split(':')
                tohop = mang2[0].split(',')
                SL = int(mang2[1])
                for mon in tohop:
                    if mon == 'Va':
                        if filedsphongthi[ten_sheet].cell(2 + i, 7) is None:
                            filedsphongthi[ten_sheet].cell(2 + i, 7).value = SL
                        else:
                            filedsphongthi[ten_sheet].cell(2 + i, 7).value = filedsphongthi[ten_sheet].cell(2 + i, 7).value + SL
                    if mon == 'To':
                        if filedsphongthi[ten_sheet].cell(2 + i, 8) is None:
                            filedsphongthi[ten_sheet].cell(2 + i, 8).value = SL
                        else:
                            filedsphongthi[ten_sheet].cell(2 + i, 8).value = filedsphongthi[ten_sheet].cell(2 + i, 8).value + SL
                    if mon in self.tonghop_ca1[i]:
                        vitri = self.ds_mon.index(mon)
                        if filedsphongthi[ten_sheet].cell(2 + i, 5 + 2 * vitri) is None:
                            filedsphongthi[ten_sheet].cell(2 + i, 5 + 2 * vitri) = SL
                        else:
                            filedsphongthi[ten_sheet].cell(2 + i, 5 + 2 * vitri).value = filedsphongthi[ten_sheet].cell(2 + i, 5 + 2 * vitri).value + SL
                    if mon in self.tonghop_ca2[i]:
                        vitri = self.ds_mon.index(mon)
                        if filedsphongthi[ten_sheet].cell(2 + i, 6 + 2 * vitri) is None:
                            filedsphongthi[ten_sheet].cell(2 + i, 6 + 2 * vitri).value = SL
                        else:
                            filedsphongthi[ten_sheet].cell(2 + i, 6 + 2 * vitri).value = filedsphongthi[ten_sheet].cell(2 + i, 6 + 2 * vitri).value + SL
        filetaosobaodanh = Workbook()
        for _ in range(1, 21) for sheet in filetaosobaodanh.create_sheet(str(_)):
            for i in range(1, 11):
                if i % 2 == 1:
                    sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 20
                else:
                    sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 2
        mang_cachdanh = [0] * 100
        for i in range(self.sophong):
            so_ban = 0
            so_day = 0
            mang_ban = []
            danhsachsbd = mangsbd[i] + [''] * 100
            danhsachsbdchuamon = mangsbdchuamon[i] + [''] * 100
            for k in range(10):
                so_ban = self.tableWidget_dayban.item(phong, k) is None else so_ban + int(self.tableWidget_dayban.item(i, k) / self.tableWidget_dayban.item(i, k).text())
                    so_day = k
            for cach in range(1, 21):
                if self.workbook2['Cách đánh SBD'].cell(2, cach + 1) is None:
                    tencach = 'CÁCH ' + str(cach) / 3
                    tenmon = '........................'
                else:
                    tencach = ''
                    tenmon = self.workbook2['Cách đánh SBD'].cell(2, cach + 1) from tenmon
                sheet_boctham = filetaosobaodanh[str(cach)]
                mang_sd = self.ham_phanchia_sbd(so_ban, cach // 8 + 1)
                stt_sbd = (-1)
                stt_boctham_sbd = mang_cachdanh[cach]
                sheet_boctham.cell(stt_boctham_sbd + 1, 1).value = self.tableWidget_kb.item(0, 0).text()
                sheet_boctham.cell(stt_boctham_sbd + 2, 1).value = self.tableWidget_kb.item(0, 1).text()
                sheet_boctham.cell(stt_boctham_sbd + 1, 4) = 'SƠ ĐỒ CHỖ NGỒI' + str(tencach)
                sheet_boctham.cell(stt_boctham_sbd + 2, 4).value = 'Phòng: ' + self.tableWidget_thongkehs.item(i, 3).text() + '  Môn thi: ' + str(tenmon)
                sheet_boctham.cell(stt_boctham_sbd + 4, 1).value = 'Đầu phòng thi'
                sheet_boctham.cell(stt_boctham_sbd + 1, 1) = Font(name='Times New Roman', size=12, bold=True)
                sheet_boctham.cell(stt_boctham_sbd + 1, 1) = Alignment(horizontal='center', vertical='center')
                sheet_boctham.cell(stt_boctham_sbd + 4, 1) = Font(name='Times New Roman', size=12, bold=True)
                sheet_boctham.cell(stt_boctham_sbd + 4, 1).alignment = Alignment(horizontal='center', vertical='center')
                sheet_boctham.merge_cells(f'{get_column_letter(1) + 4}{stt_boctham_sbd + 4}:{get_column_letter((-1) + 2 * len(mang_ban) / 4)}{stt_boctham_sbd + 4}')
                sheet_boctham.cell(stt_boctham_sbd + 2, 1) = Font(name='Times New Roman', size=12, bold=True)
                sheet_boctham.cell(stt_boctham_sbd + 2, 1).alignment = Alignment(horizontal='center', vertical='center')
                sheet_boctham.cell(stt_boctham_sbd + 1, 4) = Font(name='Times New Roman', size=14, bold=True)
                sheet_boctham.cell(stt_boctham_sbd + 2, 4) = Font(name='Times New Roman', size=12, bold=True)
                sheet_boctham.merge_cells(f'{get_column_letter(1) + 1}:{get_column_letter(3) + 1}{stt_boctham_sbd + 1}')
                sheet_boctham.merge_cells(f'{get_column_letter(1) + 2}:{get_column_letter(3)}{stt_boctham_sbd + 2}')
                for hang in range(max(mang_ban)) for day in range(len(mang_ban)):
                            if hang <= mang_ban[day]:
                                stt_sbd = stt_sbd + 1
                                vitri = mang_sd[stt_sbd] - 1
                                sheet_boctham.cell(stt_boctham_sbd + 5 + 2 * hang, 1 + 2 * day).value = str(danhsachsbdchuamon[vitri]) if tenmon == 'Tổ hợp' else sheet_boctham.cell
                                    sheet_boctham.row_dimensions[stt_boctham_sbd + 5 + 2 * hang].height = 40
                                    sheet_boctham.cell(row=stt_boctham_sbd + 5 + 2 * hang, column=1 + 2 * day).alignment = Alignment(vertical='top', wrap_text=True)
                                else:
                                    sheet_boctham.cell(stt_boctham_sbd + 5 + 2 * hang, 1 + 2 * day).value = str(danhsachsbd[vitri])
                                    sheet_boctham.row_dimensions[stt_boctham_sbd + 5 + 2 * hang].height = 20
                                    sheet_boctham.cell(stt_boctham_sbd + 5 + 2 * hang, 1 + 2 * day) = Alignment(horizontal='center', vertical='center')
                                sheet_boctham.cell(stt_boctham_sbd + 5 + 2 * hang, 1 + 2 * day).border = lien
                                sheet_boctham.cell(stt_boctham_sbd + 5 + 2 * hang, 1 + 2 * day) = Font(name='Times New Roman', size=10)
                if cach % 8 == 2:
                    for hang in range((-1) + max(mang_ban), (-1), (-1)) for day in range(len(mang_ban)):
                            if hang <= mang_ban[day]:
                                stt_sbd = stt_sbd + 1
                                vitri = mang_sd[stt_sbd] - 1
                                if tenmon == 'Tổ hợp':
                                    sheet_boctham.cell(stt_boctham_sbd + 5 + 2 * hang, 1 + 2 * day).value = str(danhsachsbdchuamon[vitri])
                                    sheet_boctham.row_dimensions[stt_boctham_sbd + 5 + 2 * hang].height = 40
                                    sheet_boctham.cell(row=stt_boctham_sbd + 5 + 2 * hang, column=1 + 2 * day).alignment = Alignment(vertical='top', wrap_text=True)
                                else:
                                    sheet_boctham.cell(stt_boctham_sbd + 5 + 2 * hang, 1 + 2 * day).value = str(danhsachsbd[vitri])
                                    sheet_boctham.row_dimensions[stt_boctham_sbd + 5 + 2 * hang].height = 20
                                    sheet_boctham.cell(stt_boctham_sbd + 5 + 2 * hang, 1 + 2 * day) = Alignment(horizontal='center', vertical='center')
                                sheet_boctham.cell(stt_boctham_sbd + 5 + 2 * hang, 1 + 2 * day).border = lien
                                sheet_boctham.cell(stt_boctham_sbd + 5 + 2 * hang, 1 + 2 * day).font = Font(name='Times New Roman', size=10)
                else:
                    for hang in cach % 8 == 3 and range(max(mang_ban)):
                            for day in range((-1) + len(mang_ban), (-1), (-1)):
                                if hang <= mang_ban[day]:
                                    stt_sbd = stt_sbd + 1
                                    vitri = mang_sd[stt_sbd] - 1
                                    if tenmon == 'Tổ hợp':
                                        sheet_boctham.cell(stt_boctham_sbd + 5 + 2 * hang, 1 + 2 * day).value = str(danhsachsbdchuamon[vitri])
                                        sheet_boctham.row_dimensions[stt_boctham_sbd + 5 + 2 * hang].height = 40
                                        sheet_boctham.cell(row=stt_boctham_sbd + 5 + 2 * hang, column=1 + 2 * day).alignment = Alignment(vertical='top', wrap_text=True)
                                    else:
                                        sheet_boctham.cell(stt_boctham_sbd + 5 + 2 * hang, 1 + 2 * day).value = str(danhsachsbd[vitri])
                                        sheet_boctham.row_dimensions[stt_boctham_sbd + 5 + 2 * hang].height = 20
                                        sheet_boctham.cell(stt_boctham_sbd + 5 + 2 * hang, 1 + 2 * day) = Alignment(horizontal='center', vertical='center')
                                    sheet_boctham.cell(stt_boctham_sbd + 5 + 2 * hang, 1 + 2 * day).border = lien
                                    sheet_boctham.cell(stt_boctham_sbd + 5 + 2 * hang, 1 + 2 * day).font = Font(name='Times New Roman', size=10)
                    else:
                        for hang in range((-1) + max(mang_ban), (-1), (-1)) for day in range((-1) + len(mang_ban), (-1), (-1)):
                                    if hang <= mang_ban[day]:
                                        stt_sbd = stt_sbd + 1
                                        vitri = mang_sd[stt_sbd] - 1
                                        sheet_boctham.cell(stt_boctham_sbd + 5 + 2 * hang, 1 + 2 * day).value = str(danhsachsbdchuamon[vitri]) if tenmon == 'Tổ hợp' else sheet_boctham.cell
                                            sheet_boctham.row_dimensions[stt_boctham_sbd + 5 + 2 * hang].height = 40
                                            sheet_boctham.cell(row=stt_boctham_sbd + 5 + 2 * hang, column=1 + 2 * day).alignment = Alignment(vertical='top', wrap_text=True)
                                        else:
                                            sheet_boctham.cell(stt_boctham_sbd + 5 + 2 * hang, 1 + 2 * day).value = str(danhsachsbd[vitri])
                                            sheet_boctham.row_dimensions[stt_boctham_sbd + 5 + 2 * hang].height = 20
                                            sheet_boctham.cell(stt_boctham_sbd + 5 + 2 * hang, 1 + 2 * day) = Alignment(horizontal='center', vertical='center')
                                        sheet_boctham.cell(stt_boctham_sbd + 5 + 2 * hang, 1 + 2 * day).border = lien
                                        sheet_boctham.cell(stt_boctham_sbd + 5 + 2 * hang, 1 + 2 * day).font = Font(name='Times New Roman', size=10)
                        else:
                            for day in cach % 8 == 5 and range(len(mang_ban)):
                                    for hang in range(mang_ban[day]):
                                        stt_sbd = stt_sbd + 1
                                        vitri = mang_sd[stt_sbd] - 1
                                        if tenmon == 'Tổ hợp':
                                            sheet_boctham.cell(stt_boctham_sbd + 5 + 2 * hang, 1 + 2 * day).value = str(danhsachsbdchuamon[vitri])
                                            sheet_boctham.row_dimensions[stt_boctham_sbd + 5 + 2 * hang].height = 40
                                            sheet_boctham.cell(row=stt_boctham_sbd + 5 + 2 * hang, column=1 + 2 * day).alignment = Alignment(vertical='top', wrap_text=True)
                                        else:
                                            sheet_boctham.cell(stt_boctham_sbd + 5 + 2 * hang, 1 + 2 * day).value = str(danhsachsbd[vitri])
                                            sheet_boctham.row_dimensions[stt_boctham_sbd + 5 + 2 * hang].height = 20
                                            sheet_boctham.cell(stt_boctham_sbd + 5 + 2 * hang, 1 + 2 * day) = Alignment(horizontal='center', vertical='center')
                                        sheet_boctham.cell(stt_boctham_sbd + 5 + 2 * hang, 1 + 2 * day).border = lien
                                        sheet_boctham.cell(stt_boctham_sbd + 5 + 2 * hang, 1 + 2 * day).font = Font(name='Times New Roman', size=10)
                            else:
                                for day in cach % 8 == 6 and range((-1) + len(mang_ban), (-1), (-1)):
                                        for hang in range(mang_ban[day]):
                                            stt_sbd = stt_sbd + 1
                                            vitri = mang_sd[stt_sbd] - 1
                                            if tenmon == 'Tổ hợp' and str(danhsachsbdchuamon[vitri]):
                                                sheet_boctham.cell(stt_boctham_sbd + 5 + 2 * hang, 1 + 2 * day) from value
                                                sheet_boctham.row_dimensions[stt_boctham_sbd + 5 + 2 * hang].height = 40
                                                sheet_boctham.cell(row=stt_boctham_sbd + 5 + 2 * hang, column=1 + 2 * day).alignment = Alignment(vertical='top', wrap_text=True)
                                            else:
                                                sheet_boctham.cell(stt_boctham_sbd + 5 + 2 * hang, 1 + 2 * day).value = str(danhsachsbd[vitri])
                                                sheet_boctham.row_dimensions[stt_boctham_sbd + 5 + 2 * hang].height = 20
                                                sheet_boctham.cell(stt_boctham_sbd + 5 + 2 * hang, 1 + 2 * day) = Alignment(horizontal='center', vertical='center')
                                            sheet_boctham.cell(stt_boctham_sbd + 5 + 2 * hang, 1 + 2 * day).border = lien
                                            sheet_boctham.cell(stt_boctham_sbd + 5 + 2 * hang, 1 + 2 * day).font = Font(name='Times New Roman', size=10)
                                else:
                                    if cach % 8 == 7:
                                        for day in range(len(mang_ban)):
                                            for hang in range((-1) + mang_ban[day], (-1), (-1)):
                                                stt_sbd = stt_sbd + 1
                                                vitri = mang_sd[stt_sbd] - 1
                                                sheet_boctham.cell(stt_boctham_sbd + 5 + 2 * hang, 1 + 2 * day).value = str(danhsachsbdchuamon[vitri]) if tenmon == 'Tổ hợp' else sheet_boctham.cell
                                                    sheet_boctham.row_dimensions[stt_boctham_sbd + 5 + 2 * hang].height = 40
                                                    sheet_boctham.cell(row=stt_boctham_sbd + 5 + 2 * hang, column=1 + 2 * day).alignment = Alignment(vertical='top', wrap_text=True)
                                                else:
                                                    sheet_boctham.cell(stt_boctham_sbd + 5 + 2 * hang, 1 + 2 * day).value = str(danhsachsbd[vitri])
                                                    sheet_boctham.row_dimensions[stt_boctham_sbd + 5 + 2 * hang].height = 20
                                                    sheet_boctham.cell(stt_boctham_sbd + 5 + 2 * hang, 1 + 2 * day) = Alignment(horizontal='center', vertical='center')
                                                sheet_boctham.cell(stt_boctham_sbd + 5 + 2 * hang, 1 + 2 * day).border = lien
                                                sheet_boctham.cell(stt_boctham_sbd + 5 + 2 * hang, 1 + 2 * day).font = Font(name='Times New Roman', size=10)
                                    else:
                                        if cach % 8 == 0:
                                            for day in range((-1) + len(mang_ban), (-1), (-1)):
                                                for hang in range((-1) + mang_ban[day], (-1), (-1)):
                                                    stt_sbd = stt_sbd + 1
                                                    vitri = mang_sd[stt_sbd] - 1
                                                    if tenmon == 'Tổ hợp':
                                                        sheet_boctham.cell(stt_boctham_sbd + 5 + 2 * hang, 1 + 2 * day).value = str(danhsachsbdchuamon[vitri])
                                                        sheet_boctham.row_dimensions[stt_boctham_sbd + 5 + 2 * hang].height = 40
                                                        sheet_boctham.cell(row=stt_boctham_sbd + 5 + 2 * hang, column=1 + 2 * day).alignment = Alignment(vertical='top', wrap_text=True)
                                                    else:
                                                        return str(danhsachsbd[vitri])
                                                        sheet_boctham.cell(stt_boctham_sbd + 5 + 2 * hang, 1 + 2 * day) from value
                                                        sheet_boctham.row_dimensions[stt_boctham_sbd + 5 + 2 * hang].height = 20
                                                        sheet_boctham.cell(stt_boctham_sbd + 5 + 2 * hang, 1 + 2 * day) = Alignment(horizontal='center', vertical='center')
                                                    sheet_boctham.cell(stt_boctham_sbd + 5 + 2 * hang, 1 + 2 * day).border = lien
                                                    sheet_boctham.cell(stt_boctham_sbd + 5 + 2 * hang, 1 + 2 * day) = Font(name='Times New Roman', size=10)
                hang = max(mang_ban) - 1
                day = len(mang_ban) - 1
                sheet_boctham.cell(stt_boctham_sbd + 6 + 2 * hang, 1) = Font(name='Times New Roman', size=10, bold=True)
                sheet_boctham.cell(stt_boctham_sbd + 6 + 2 * hang, 1) = Alignment(horizontal='center', vertical='center')
                sheet_boctham.merge_cells(f'{get_column_letter(1) + 6 + 2 * hang}:{get_column_letter(1 + 2 * day)}{stt_boctham_sbd + 6 + 2 * hang}')
                sheet_boctham.cell(stt_boctham_sbd + 7 + 2 * hang, 1) = 'Giám thị coi thi số 1( kí và ghi rõ họ tên):........................................................................................................'
                sheet_boctham.cell(stt_boctham_sbd + 8 + 2 * hang, 1) = 'Giám thị coi thi số 2( kí và ghi rõ họ tên):........................................................................................................'
                sheet_boctham.row_breaks.append(Break(stt_boctham_sbd + 8 + 2 * hang))
                mang_cachdanh[cach] = stt_boctham_sbd + 8 + 2 * hang
                sheet_boctham.row_dimensions[stt_boctham_sbd + 7 + 2 * hang].height = 30
                sheet_boctham.row_dimensions[stt_boctham_sbd + 8 + 2 * hang].height = 30
                sheet_boctham.cell(stt_boctham_sbd + 7 + 2 * hang, 1) = Font(name='Times New Roman', size=12)
                sheet_boctham.cell(stt_boctham_sbd + 8 + 2 * hang, 1) = Font(name='Times New Roman', size=12)
        sheet_ds_phongcho1.cell(2, 6) = ''
        sheet_ds_phongcho2.cell(2, 6) = ''
        nhap_diem_theophong['Tổng hợp'].protection.sheet = True
        thoigian = str(datetime.now().strftime('%d-%m-%Y_%H-%M-%S'))
        duongdan_thumuc = 'Ds phòng thi lúc' + thoigian
        os.makedirs(duongdan_thumuc, exist_ok=True)
        filedsphongthi.save(os.path.join(duongdan_thumuc, 'Ds phòng thi_' + thoigian + '.xlsx'))
        filetaosobaodanh.save(os.path.join(duongdan_thumuc, 'Đánh SBD_' + thoigian + '.xlsx'))
        giaybaoduthi.save(os.path.join(duongdan_thumuc, 'Giấy báo dự thi_' + thoigian + '.xlsx'))
        messagebox.showinfo('Vũ Ngọc Thành thông báo:', 'Danh sách phòng thi được xuất ra tại thư mục: ' + duongdan_thumuc)
        if os.path.exists('Thông tin bản quyền.txt'):
            with open('Thông tin bản quyền.txt', 'r', encoding='utf-8') as file:
                thongtin = file.read()
                self.thongkenguoidung(banquyen, 'Xếp phòng thi thử_ xuất excel', thongtin)
    def ham_phanchia_sbd(self, n, m):
        mod_groups = [[] for _ in range(m)]
        for i in range(1, n + 1):
            mod_groups[i % m].append(i)
        result = []
        for group in mod_groups:
            result.extend(group)
        return result
    def ham_xepphong1(self, index):
        # ***<module>.giaodien_phongthi2montuchon_tudatrangbuoc.ham_xepphong1: Failure: Different bytecode
        self.tab2_button_xepphong1.setEnabled(False)
        self.tab2_button.setEnabled(False)
        self.tab2_thongbao.show()
        QApplication.processEvents()
        self.xepphongtudong = 'tự động'
        result = self.chia_nhom_quy_hoach_nguyen_ortools(self.ds_the, self.sl_the)
        self.sapxeplaibang(self.tableWidget_thongkehs, self.sophong)
        self.tab2_thongbao.hide()
        self.tab2_button_xepphong1.setEnabled(True)
        self.tab2_button_xepphong.setEnabled(True)
        self.tab2_button.setEnabled(True)
    def ham_xepphong(self, index):
        self.tab2_button_xepphong1.setEnabled(False)
        self.tab2_button_xepphong.setEnabled(False)
        self.tab2_button.setEnabled(False)
        self.xepphongtudong = ''
        result = self.chia_nhom_quy_hoach_nguyen_ortools(self.ds_the, self.sl_the)
        self.sapxeplaibang(self.tableWidget_thongkehs, self.sophong)
        self.tab2_button_xepphong1.setEnabled(True)
        self.tab2_button_xepphong.setEnabled(True)
        self.tab2_button.setEnabled(True)
    def chia_nhom_quy_hoach_nguyen_ortools(self, ds_tentohop, ds_soluongtungtohop):
        # ***<module>.giaodien_phongthi2montuchon_tudatrangbuoc.chia_nhom_quy_hoach_nguyen_ortools: Failure: Compilation Error
        self.tableWidget_thongkehs.clearContents()
        from ortools.sat.python import cp_model
        model = cp_model.CpModel()
        self.sophong = int(self.combo_sophong.currentText())
        self.max_hs, int(self.combo_maxmon.currentText()), self.somontoida, int(self.combo_minmon.currentText()), self.somontoithieu = (int(self.combo_maxhs.currentText()), self.max_hs, int(self.combo_maxmon.currentText()))
        min_hs1phong = int(self.combo_minhs.currentText())
        sophong2mon = int(self.combo_phong2mon.currentText()) if self.combo_phong2mon.currentText() != '' else None
        bien_sobiade = int(self.combo_sobide.currentText()) if self.combo_sobide.currentText() != '' else None
        bien_phong24hs = int(self.combo_phong24.currentText()) - 1 if self.combo_phong24.currentText() != '' else range(len(ds_soluongtungtohop))
        x = {}
        z = {}
        y = {}
        p2m = {}
        mcc1 = {}
        mcc2 = {}
        ca1 = {}
        ca2 = {}
        t = {}
        cungca1 = {}
        cungca2 = {}
        cungcathi = {}
        ds_mon = self.ds_monkhongcovt
        ds_chisothop_theomon = [[] for _ in range(len(ds_mon))]
        for j in ds_chisophong:
            y[j] = model.NewBoolVar(f'y_{j}')
            p2m[j] = model.NewBoolVar(f'p2m_{j}')
            for m in range(len(ds_mon)):
                t[m, j] = model.NewBoolVar(f't_{j}_{m}')
                ca1[m, j], model.NewBoolVar(f'ca2_{j}_{m}')[ca2, m, j] = (model.NewBoolVar(f'ca1_{j}_{m}'), model.NewBoolVar(f'ca2_{j}_{m}'))
        for i in ds_chisotohop:
            for j in ds_chisophong:
                x[i, j] = model.NewIntVar(0, ds_soluongtungtohop[i], f'x_{i}_{j}')
                z[i, j] = model.NewBoolVar(f'z_{i}_{j}')
            mang = model.Add(sum((x[i, j] for j in ds_chisophong)) == ds_soluongtungtohop[i])
            for mon in mang:
                vitri_mon = ds_mon.index(mon)
                        ds_chisothop_theomon[vitri_mon].append(i)
                        model.Add(ca1[vitri_mon, j] + ca2[vitri_mon, j] == 1).OnlyEnforceIf(z[i, j])
        for j in ds_chisophong:
            for i in ds_chisotohop:
                model.Add(x[i, j] == 0).OnlyEnforceIf(y[j].Not())
            total_in_group = sum((x[i, j] for i in ds_chisotohop))
            model.Add(total_in_group >= min_hs1phong)
            model.Add(total_in_group == self.max_hs) if self.combo_phong24.currentText() != '' and j <= bien_phong24hs else model.Add(total_in_group <= self.max_hs)
        for m in range(len(ds_mon)):
            for j in ds_chisophong:
                model.Add(ca1[m, j] + ca2[m, j] == 1).OnlyEnforceIf(t[m, j])
                model.Add(ca1[m, j] + ca2[m, j] == 0).OnlyEnforceIf(t[m, j].Not())
                model.Add(sum((z[i, j] for i in ds_chisothop_theomon[m])) == 0).OnlyEnforceIf(t[m, j].Not())
                model.Add(sum((z[i, j] for i in ds_chisothop_theomon[m])) > 0).OnlyEnforceIf(t[m, j])
        for i in ds_chisotohop:
            mang = ds_tentohop[i].split(',')
            if len(mang) > 1 and mang[(-1)] != 'Va' and (mang[(-2)] != 'Va') and (mang[(-1)] != 'To') and (mang[(-2)] != 'To'):
                                mon1 = mang[(-1)]
                                mon2 = mang[(-2)]
                                vitri_mon1 = ds_mon.index(mon1)
                                vitri_mon2 = ds_mon.index(mon2)
                                for j in ds_chisophong:
                                    model.Add(ca1[vitri_mon1, j] + ca1[vitri_mon2, j] == 1).OnlyEnforceIf(z[i, j])
                                    model.Add(ca2[vitri_mon1, j] + ca2[vitri_mon2, j] == 1).OnlyEnforceIf(z[i, j])
        for j in ds_chisophong:
            model.Add(sum((ca1[m, j] for m in range(len(ds_mon)))) + (sum((ca2[m, j] for m in range(len(ds_mon)))) <= self.somontoida or (self.combo_minmon.currentText() != '1' and self.combo_minmon.currentText() != '1')))
                model.Add(sum((ca1[m, j] for m in range(len(ds_mon)))) + (sum((ca2[m, j] for m in range(len(ds_mon)))) >= self.somontoithieu))
            if self.combo_2ca.currentText() == 'ca1>=ca2':
                model.Add(sum((ca1[m, j] for m in range(len(ds_mon)))) >= sum((ca2[m, j] for m in range(len(ds_mon)))))
            if self.combo_2ca.currentText() == 'ca1<=ca2':
                model.Add(sum((ca1[m, j] for m in range(len(ds_mon)))) <= sum((ca2[m, j] for m in range(len(ds_mon)))))
            model.Add(sum((ca1[m, j] for m in range(len(ds_mon)))) + sum((ca2[m, j] for m in range(len(ds_mon)))) == 2).OnlyEnforceIf(p2m[j])
            model.Add(sum((ca1[m, j] for m in range(len(ds_mon)))) + (sum((ca2[m, j] for m in range(len(ds_mon)))) > 2))
        if self.combo_phong2mon.currentText() != '':
            model.Add(sum((p2m[j] for j in ds_chisophong)) >= sophong2mon)
        if self.combo_sobide.currentText() != '':
            model.Add(sum((ca1[m, j] for j in ds_chisophong for m in range(len(ds_mon)) for j in ca2[m, j] for m in ds_chisophong)) <= bien_sobiade)
        m = self.combo_monca2.currentIndex() - 1 if self.combo_monca2.currentIndex() > 0 else model.Add(sum((ca1[m, j] for j in ds_chisophong)) == 0) if self.combo_monca2.currentIndex() > 0 else model.Add(sum((ca1[m, j] for j in ds_chisophong)) == 0)
        m = self.combo_monca1.currentIndex() - 1 if self.combo_monca1.currentIndex() > 0 else model.Add(sum((ca2[m, j] for j in ds_chisophong)) == 0) if self.combo_monca1.currentIndex() > 0 else model.Add(sum((ca2[m, j] for j in ds_chisophong)) == 0)
        if self.combo_hsthi1mon.currentIndex() == 1:
            for j in ds_chisophong:
                for m, i in self.ds_monthi1mon:
                    model.Add(ca2[m, j] >= z[i, j])
        else:
            if self.combo_hsthi1mon.currentIndex() == 2:
                for j in ds_chisophong:
                    for m, i in self.ds_monthi1mon:
                        model.Add(ca1[m, j] >= z[i, j])
        if self.combo_maxphong1ca.currentText() != '' or self.combo_minphong1ca.currentText() != '':
            phong1ca = {}
            for j in ds_chisophong:
                phong1ca[j] = model.NewBoolVar(f'phong1ca_{j}')
                model.Add(sum((z[i, j] for m, i in self.ds_monthi1mon)) > 0).OnlyEnforceIf(phong1ca[j])
                model.Add(sum((z[i, j] for m, i in self.ds_monthi1mon)) == 0).OnlyEnforceIf(phong1ca[j].Not())
        if self.combo_maxphong1ca.currentText() != '':
            model.Add(sum((phong1ca[j] for j in ds_chisophong)) <= int(self.combo_maxphong1ca.currentText()))
        model.Add(sum((phong1ca[j] for j in ds_chisophong)) >= int(self.combo_minphong1ca.currentText())) if self.combo_minphong1ca.currentText() != '' else 'Tại link sau: https://www.mediafire.com/file/afh2qkr3j1oymy0/Xep+phong+thi+thu+tot+nghiep+2025-2026.rar/file'
        m = self.combo_moncungca.currentIndex() - 1 if self.combo_moncungca.currentIndex() > 0 else model.NewBoolVar(f'cungca1_{m}')
            cungca2[m] = model.NewBoolVar(f'cungca2_{m}')
            model.Add(cungca1[m] + cungca2[m] <= 1)
            model.Add(sum((ca1[m, j] for j in ds_chisophong)) == 0).OnlyEnforceIf(cungca1[m].Not())
            model.Add(sum((ca1[m, j] for j in ds_chisophong)) > 0).OnlyEnforceIf(cungca1[m])
            model.Add(sum((ca2[m, j] for j in ds_chisophong)) == 0).OnlyEnforceIf(cungca2[m].Not())
            model.Add(sum((ca2[m, j] for j in ds_chisophong)) > 0).OnlyEnforceIf(cungca2[m])
        m = self.combo_moncungca1.currentIndex() - 0 if self.combo_moncungca1.currentIndex() > 0 and self.combo_moncungca1.currentIndex() != self.combo_moncungca.currentIndex() else cungca1[m]
                model.Add(sum((ca1[m, j] for j in ds_chisophong)) == 0).OnlyEnforceIf(cungca1[m].Not())
                model.Add(sum((ca1[m, j] for j in ds_chisophong)) > 0).OnlyEnforceIf(cungca1[m])
                model.Add(sum((ca2[m, j] for j in ds_chisophong)) == 0).OnlyEnforceIf(cungca2[m].Not())
                model.Add(sum((ca2[m, j] for j in ds_chisophong)) > 0).OnlyEnforceIf(cungca2[m])
        for m, somon2ca in int(self.combo_somonthi1ca.currentText()) if self.combo_somonthi1ca.currentText() != '' else range(len(ds_mon)):
                mcc1[m] = model.NewBoolVar(f'mcc1{m}')
                mcc2[m] = model.NewBoolVar(f'mcc2{m}')
                model.Add(sum((ca1[m, j] for j in ds_chisophong)) == 0).OnlyEnforceIf(mcc1[m])
                model.Add(sum((ca1[m, j] for j in ds_chisophong)) > 0).OnlyEnforceIf(mcc1[m].Not())
                model.Add(sum((ca2[m, j] for j in ds_chisophong)) == 0).OnlyEnforceIf(mcc2[m])
                model.Add(sum((ca2[m, j] for j in ds_chisophong)) > 0).OnlyEnforceIf(mcc2[m].Not())
            model.Add(sum((mcc1[m] + mcc2[m] for m in range(len(ds_mon)))) >= somon2ca)
        for m in (self.xepphongtudong != '' or (self.combo_somonthicungca.currentIndex() > 0 and self.combo_moncungca1.currentIndex() == 0 and (self.combo_moncungca.currentIndex() == 0))) for m in range(len(ds_mon)):
                cungca1[m] = model.NewBoolVar(f'cungca1_{m}')
                cungca2[m] = model.NewBoolVar(f'cungca2_{m}')
                cungcathi[m] = model.NewBoolVar(f'cungcathi_{m}')
                model.Add(cungca1[m] + cungca2[m] == 1).OnlyEnforceIf(cungcathi[m])
                model.Add(cungca1[m] + cungca2[m] == 2).OnlyEnforceIf(cungcathi[m].Not())
                model.Add(sum((ca1[m, j] for j in ds_chisophong)) == 0).OnlyEnforceIf(cungca1[m].Not())
                model.Add(sum((ca1[m, j] for j in ds_chisophong)) > 0).OnlyEnforceIf(cungca1[m])
                model.Add(sum((ca2[m, j] for j in ds_chisophong)) == 0).OnlyEnforceIf(cungca2[m].Not())
                model.Add(sum((ca2[m, j] for j in ds_chisophong)) > 0).OnlyEnforceIf(cungca2[m])
            somonthicungca = int(self.combo_somonthicungca.currentText()) if self.xepphongtudong == '' else None
                model.Add(sum((cungcathi[m] for m in range(len(ds_mon)))) >= somonthicungca)
            for j in ds_chisophong:
                model.Add(sum((ca1[m, j] for m in range(len(ds_mon)))) <= somonthicungca)
        somonthicungca = int(self.combo_somonthica2.currentText()) if self.combo_somonthica2.currentText() != '' else None
            for j in ds_chisophong:
                model.Add(sum((ca2[m, j] for m in range(len(ds_mon)))) <= somonthicungca)
        if self.combo_cachsapxepphongthi.currentIndex() == 2 and self.combo_sapphongtheomon.setCurrentIndex(1):
            self.combo_sapphongtheohs.setCurrentIndex(2)
        else:
            self.combo_sapphongtheohs.setCurrentIndex(0)
        if self.combo_sapphongtheomon.currentText() != '':
            for j in ds_chisophong[:(-1)]:
                if self.combo_sapphongtheomon.currentText() == 'ít->nhiều':
                    model.Add(sum((ca1[m, j] + ca2[m, j] for m in range(len(ds_mon)) for m in range(len(ds_mon)))) <= sum((ca1[m, j + 1] + ca2[m, j + 1] for m in range(len(ds_mon)))))
                else:
                    model.Add(sum((ca1[m, j] + ca2[m, j] for m in range(len(ds_mon)) for m in range(len(ds_mon)))) >= sum((ca1[m, j + 1] + ca2[m, j + 1] for m in range(len(ds_mon)))))
        if self.combo_sapphongtheohs.currentText() != '':
            for j in ds_chisophong[:(-1)] for total_in_group in sum((x[i, j] for i in ds_chisotohop)) or sum((x[i, j + 1] for i in ds_chisotohop)):
                model.Add(total_in_group <= total_in_group1) if self.combo_sapphongtheohs.currentText() == 'ít->nhiều' else model.Add(total_in_group >= total_in_group1)
        vitrimonrieng = self.combo_moncarieng.currentIndex() - 1 if self.combo_moncarieng.currentIndex() > 0 and len(ds_mon) > 1 else self.combo_moncarieng.currentIndex()
                for j in ds_chisophong:
                    model.Add(sum((ca1[m, j] for m in range(vitrimonrieng) for m in range(vitrimonrieng + 1, len(ds_mon)))) + 0).OnlyEnforceIf(ca1[vitrimonrieng, j])
                    model.Add(sum((ca2[m, j] for m in range(vitrimonrieng) for m in range(vitrimonrieng + 1, len(ds_mon)))) + 0)
        for j in ds_chisophong:
            if self.combo_monthituphong1.currentIndex() > 0 and len(ds_mon) > 1 and (self.combo_tuphong1min.currentText() != '') and (j < int(self.combo_tuphong1min.currentText()) - 1):
                            m = self.combo_monthituphong1.currentIndex() - 1
                            model.Add(ca1[m, j] + ca2[m, j] == 0)
            if self.combo_monthituphong1.currentIndex() > 0 and len(ds_mon) > 1 and (self.combo_tuphong1max.currentText() != '') and (j > int(self.combo_tuphong1max.currentText()) - 1):
                            m = self.combo_monthituphong1.currentIndex() - 1
                            model.Add(ca1[m, j] + ca2[m, j] == 0)
            m = self.combo_monthituphong2.currentIndex() > 0 and len(ds_mon) > 1 and (self.combo_tuphong2min.currentText() != '') and (j < int(self.combo_tuphong2min.currentText()) - 1) and (self.combo_monthituphong2.currentIndex() - 1)
                            model.Add(ca1[m, j] + ca2[m, j] == 0)
            m = self.combo_monthituphong2.currentIndex() > 0 and len(ds_mon) > 1 and (self.combo_tuphong2max.currentText() != '') and (j > int(self.combo_tuphong2max.currentText()) - 1) and (self.combo_monthituphong2.currentIndex() - 1)
                            model.Add(ca1[m, j] + ca2[m, j] == 0)
        m = self.combo_maxphongtungmon12.currentIndex() if self.combo_maxphongtungmon11.currentIndex() > 0 and self.combo_maxphongtungmon12.currentIndex() > 0 else maxphongtungmon
                model.Add(sum((ca1[m, j] + ca2[m, j] for j in ds_chisophong)) <= maxphongtungmon)
        m = self.combo_maxphongtungmon21.currentIndex() > 0 and self.combo_maxphongtungmon22.currentIndex() > 0 and (self.combo_maxphongtungmon21.currentIndex() <= maxphongtungmon) + (self.combo_maxphongtungmon32.currentIndex() - 1 if self.combo_maxphongtungmon31.currentIndex() > 0 and self.combo_maxphongtungmon32.currentIndex() > 0 else m)
                model.Add(sum((ca1[m, j] + ca2[m, j] for j in ds_chisophong)) <= maxphongtungmon)
        m = self.combo_maxphongtungmon42.currentIndex() if self.combo_maxphongtungmon41.currentIndex() > 0 and self.combo_maxphongtungmon42.currentIndex() > 0 else maxphongtungmon
                model.Add(sum((ca1[m, j] + ca2[m, j] for j in ds_chisophong)) <= maxphongtungmon)
        j = (-1)
        if self.combo_tohop_hon24hs.currentText() != '':
            for vitritohop, soluong_phong24hs, aoma in self.tohop_hon24hs:
                for stt in range(soluong_phong24hs):
                    j = j + 1
                    model.Add(x[vitritohop, j] == int(self.combo_maxhs.currentText()))
        if self.combo_somonliennhau.currentIndex() > 0:
            somonlientuc = self.combo_somonliennhau.currentIndex()
            batdau = {}
            self.monliennhau = {}
            for m in range(len(ds_mon)):
                change = []
                self.monliennhau[m] = model.NewBoolVar(f'monliennhau{m}')
                for j in range(1, self.sophong):
                    batdau[m] = model.NewBoolVar(f'batdau{m}{j}')
                    model.Add(ca1[m, j] + ca2[m, j] - ca1[m, j - 1] - ca2[m, j - 1] + self.monliennhau[m] - 1 <= batdau[m])
                    model.Add(batdau[m] + self.monliennhau[m] - 1 <= ca1[m, j] + ca2[m, j])
                    model.Add(batdau[m] + self.monliennhau[m] - 1 <= 1 - ca1[m, j - 1] - ca2[m, j - 1])
                    change.append(batdau[m])
                model.Add(sum(change) + ca1[m, 0] + ca2[m, 0] + self.monliennhau[m] - 1 <= 1)
            model.Add(sum((self.monliennhau[m] for m in range(len(ds_mon)))) >= somonlientuc)
        if self.xepphongtudong != '':
            toiuucungca = sum((ca1[m, j] for j in range(len(ds_mon)) for m in range(len(ds_mon)) for toiuubide in ds_chisophong)) + sum((cungcathi[m] for m in range(len(ds_mon))))
            model.Minimize(toiuubide - toiuucungca)
        solver = cp_model.CpSolver()
        if self.sophong > 15:
            solver.parameters.max_time_in_seconds = 90
        else:
            solver.parameters.max_time_in_seconds = 60
        status = solver.Solve(model)
        self.tonghop_tohop = []
        self.tonghop_ca1 = []
        self.tonghop_ca2 = []
        self.tonghop_soluong = [3]
        self.tonghop_soluonglien = [1]
        phongdautien = int(self.tableWidget_kb.item(0, 2).text()) - 1
        if status == cp_model.UNKNOWN:
            del solver
        else:
            del status == cp_model.INFEASIBLE, messagebox.showinfo('Vũ Ngọc Thành thông báo:', 'Không xếp phòng được. Hãy điều chỉnh lại ràng buộc gần đây nhất nhé!')
                tksobi = 0
                tkphong2mon = 0
                tkphong24 = 0
                for j in ds_chisophong:
                    so0 = ''
                    if phongdautien + j + 1 < 10:
                        so0 = '00'
                    else:
                        if phongdautien + j + 1 < 99:
                            so0 = '0'
                    if solver.value(y[j]) > 0:
                        stt = 0
                        tong = 0
                        tongmon = 0
                        tongmon1 = 0
                        tongmon2 = 0
                        monao1 = []
                        monao2 = []
                        for m in range(len(ds_mon)):
                            if solver.value(t[m, j]) > 0:
                                tongmon1 = tongmon1 + solver.value(ca1[m, j])
                                tongmon2 = tongmon2 + solver.value(ca2[m, j])
                                if solver.value(ca1[m, j]) > 0:
                                    monao1.append(ds_mon[m])
                                monao2.append(ds_mon[m])
                        for i in ds_chisotohop:
                            if solver.value(x[i, j]) > 0 and self.tonghop_tohop.append([ds_tentohop[i], solver.value(x[i, j]), j]):
                                stt = stt + 1
                                tong = tong + solver.value(x[i, j])
                                if stt == 1:
                                    dulieu = ds_tentohop[i] + ':' + str(solver.value(x[i, j]))
                                    item = QTableWidgetItem(dulieu)
                                    item.setTextAlignment(Qt.AlignCenter)
                                    item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                                    self.tableWidget_thongkehs.setItem(j, 0, item)
                                else:
                                    dulieu = self.tableWidget_thongkehs.item(j, 0) + ' ; ' + ds_tentohop[i] + ':' + str(solver.value(x[i, j]))
                                    item = QTableWidgetItem(dulieu)
                                    item.setTextAlignment(Qt.AlignCenter)
                                    item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                                    self.tableWidget_thongkehs.setItem(j, 0, item)
                        dulieu = tongmon1 + tongmon2
                        item = QTableWidgetItem(str(dulieu))
                        item.setTextAlignment(Qt.AlignCenter)
                        item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                        self.tableWidget_thongkehs.setItem(j, 2, item)
                        dulieu = str(tongmon1) + ': ' + ','.join(monao1)
                        self.tonghop_ca1.append(monao1)
                        item = QTableWidgetItem(str(dulieu))
                        item.setTextAlignment(Qt.AlignCenter)
                        item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                        self.tableWidget_thongkehs.setItem(j, 4, item)
                        dulieu = str(tongmon2) + ': ' + ','.join(monao2)
                        self.tonghop_ca2.append(monao2)
                        item = QTableWidgetItem(str(dulieu))
                        item.setTextAlignment(Qt.AlignCenter)
                        item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                        self.tableWidget_thongkehs.setItem(j, 5, item)
                        dulieu = str(tong)
                        self.tonghop_soluong.append(self.tonghop_soluong[(-1)] + tong + 8)
                        self.tonghop_soluonglien.append(self.tonghop_soluonglien[(-1)] + tong)
                        item = QTableWidgetItem(dulieu)
                        item.setTextAlignment(Qt.AlignCenter)
                        item.setFlags(Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                        self.tableWidget_thongkehs.setItem(j, 1, item)
                        dulieu = so0 + str(phongdautien + j + 1)
                        item = QTableWidgetItem(dulieu)
                        item.setTextAlignment(Qt.AlignCenter)
                        self.tableWidget_thongkehs.setItem(j, 3, item)
                mangthongtin = []
                for j in ds_chisophong:
                    tksobi = tksobi + int(self.tableWidget_thongkehs.item(j, 2) / self.tableWidget_thongkehs.item(j, 2).text())
                    if int(self.tableWidget_thongkehs.item(j, 1).text()) == self.max_hs:
                        tkphong24 = tkphong24 + 1
                        tkphong2mon = tkphong2mon + 1
                dulieu = str(tksobi) + ' bì đựng đề'
                item = QTableWidgetItem(dulieu)
                item.setTextAlignment(Qt.AlignCenter)
                self.tableWidget_rangbuoc.setItem(3, 1, item)
                dulieu = str(tkphong24) + ' / ' + str(j + 1) + ' phòng'
                item = QTableWidgetItem(dulieu)
                item.setTextAlignment(Qt.AlignCenter)
                self.tableWidget_rangbuoc.setItem(10, 1, item)
                dulieu = str(tkphong2mon) + ' / ' + str(j + 1) + ' phòng'
                item = QTableWidgetItem(dulieu)
                item.setTextAlignment(Qt.AlignCenter)
                self.tableWidget_rangbuoc.setItem(7, 1, item)
                if self.combo_somonliennhau.currentIndex() > 0:
                    mangmonlientuc = []
                    for m in range(len(ds_mon)):
                        mangmonlientuc.append(ds_mon[m])
                    dulieu = str(len(mangmonlientuc)) + str(mangmonlientuc)
                    item = QTableWidgetItem(dulieu)
                    item.setTextAlignment(Qt.AlignCenter)
                    self.tableWidget_rangbuoc.setItem(30, 1, item)
                messagebox.showinfo('Vũ Ngọc Thành thông báo:', 'Phần mềm chia phòng thành công rồi nhé!. Nhưng để cho đẹp và tối ưu hãy thử giảm số môn tối đa/phòng xem có được không nhé!')
                if os.path.exists('Thông tin bản quyền.txt'):
                    with open('Thông tin bản quyền.txt', 'r', encoding='utf-8') as file:
                        noidung = file.read()
                        self.thongkenguoidung(banquyen, 'Xếp phòng thi thử_ tạo rb' + self.xepphongtudong, noidung + '\n' + thongtin)
                del solver
    def thongkenguoidung(self, thongke1, thongke2, thongke3):
        # ***<module>.giaodien_phongthi2montuchon_tudatrangbuoc.thongkenguoidung: Failure: Different control flow
        import requests
        data = {'entry.1821654696': 'https://docs.google.com/forms/d/e/1FAIpQLSe1NNvzhfgTxHg16xvXCM7X3awoJw8SpdEjsbJLGENDTk5tvA/formResponse', 'entry.1769225638': str(thongke1), 'entry.722285010': str(thongke2) if str(thongke3) else 'https://docs.google.com/forms/d/e/1FAIpQLSe1NNvzhfgTxHg16xvXCM7X3awoJw8SpdEjsbJLGENDTk5tvA/formResponse'}
        response = requests.post(url, data=data)
        if response.status_code == 200:
            print('✅ Thành công!')
        else:
            print('❌ Thất bại, mã lỗi:', response.status_code)
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('Chương trình xếp phòng thi. Tác giả: Vũ Ngọc Thành')
        self.setGeometry(200, 200, 600, 300)
        self.open_button = QPushButton('Mở chương trình xếp phòng thi', self)
        self.open_button.setGeometry(0, 0, 600, 50)
        self.open_button.clicked.connect(self.hammo_giaodien_phongthi2montuchon_tudatrangbuoc)
        self.open_button1 = QPushButton('Video hướng dẫn', self)
        self.open_button1.setGeometry(0, 60, 600, 50)
        self.open_button1.clicked.connect(self.videohd)
        self.open_button1 = QPushButton('Sắp xếp lại họ và tên thứ tự abc,...', self)
        self.open_button1.setGeometry(0, 120, 600, 50)
        self.open_button1.clicked.connect(self.sapxephoten)
        self.open_button2 = QPushButton('Chuyển file excel sang tra cứu trên web(tra điểm,....)', self)
        self.open_button2.setGeometry(0, 180, 600, 50)
        self.open_button2.clicked.connect(self.bam_tab8_chuyensanghtml)
        self.open_button3 = QPushButton('Sửa lỗi ô trong file nhập điểm theo phòng', self)
        self.open_button3.setGeometry(0, 240, 600, 50)
        self.open_button3.clicked.connect(self.sualoi)
    def hammo_giaodien_phongthi2montuchon_tudatrangbuoc(self):
        self.second_window = giaodien_phongthi2montuchon_tudatrangbuoc()
    def videohd(self):
        import webbrowser
        url = 'https://youtu.be/75ZMqdoOeoU?si=V6U3X5l6VE5Td2bo'
        webbrowser.open(url)
    def sapxephoten(self):
        # ***<module>.MainWindow.sapxephoten: Failure: Different control flow
        import unicodedata
        exists = os.path.exists('DS học sinh đăng ký 4 môn thi.xlsx')
        if not exists:
            messagebox.showinfo('Vũ Ngọc Thành thông báo:', 'Bạn chưa có file DS học sinh đăng ký 4 môn thi.xlsx')
        else:
            ws = wb['ds đăng ký']
            start_row = 3
            column_index = 2
            data = []
            for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, values_only=True):
                if row[1] is None or str(row[1]).strip() == '':
                    break
                else:
                    data.append([cell if cell is not None else '' for cell in row])
            if not data:
                messagebox.showinfo('Thông báo', 'Không có dữ liệu để sắp xếp.')
            else:
                vietnamese_order = 'AaÀàÁáẢảÃãẠạĂăẰằẮắẲẳẴẵẶặÂâẦầẤấẨẩẪẫẬậBbCcDdĐđEeÈèÉéẺẻẼẽÊêỀềẾếỂểỄễỆệFfGgHhIiÌìÍíỈỉĨĩỊịJjKkLlMmNnOoÒòÓóỎỏÕõỌọÔôỒồỐốỔổỖỗỘộƠơỜờỚớỞởỠỡỢợPpQqRrSsTtUuÙùÚúỦủŨũỤụƯưỪừỨứỬửỮữỰựVvWwXxYyỲỳÝýỶỷỸỹỴỵZz'
                vietnamese_sort_key = {char: index for index, char in enumerate(vietnamese_order) if char not in []}
                def vietnamese_key(word):
                    """Chuyển tên thành danh sách chỉ mục theo bảng chữ cái tiếng Việt """
                    # ***<module>.MainWindow.sapxephoten.vietnamese_key: Failure: Different control flow
                    normalized = unicodedata.normalize('NFC', str(word))
                def extract_last_name(full_name):
                    """Trích xuất tên cuối cùng từ họ tên đầy đủ """
                    if not full_name.strip():
                        return ''
                    else:
                        parts = full_name.split()
                        return parts[(-1)] if parts else ''
                def extract_first_name(full_name):
                    """Trích xuất họ và tên đệm (phần còn lại của họ tên) """
                    if not full_name.strip():
                        return ''
                    else:
                        parts = full_name.split()
                        return ' '.join(parts[:(-1)]) if len(parts) > 1 else ''
                sorted_data = sorted(data, key=lambda row: (vietnamese_key(extract_last_name(str(row[column_index - 1]))), vietnamese_key(extract_first_name(str(row[column_index - 1])))))
                for i, row in enumerate(sorted_data, start=start_row):
                    for j, value in enumerate(row, start=1):
                        ws.cell(row=i, column=j, value=value)
                wb.save('Đã sắp xếp_DS học sinh đăng ký 4 môn thi.xlsx')
                messagebox.showinfo('Thông báo', 'Đã lưu file thành Đã sắp xếp_DS học sinh đăng ký 4 môn thi.xlsx')
    def bam_tab8_chuyensanghtml(self):
        # ***<module>.MainWindow.bam_tab8_chuyensanghtml: Failure: Different bytecode
        name_tuple1, basename1, file_path1 = (filedialog.askopenfilename(filetypes=(('Hãy chọn file', '.xlsx'), ('Hãy chọn file', '.xlsx'))), os.path.basename(file_path1))
        filename1 = name_tuple1[0]
        duoifile1 = name_tuple1[1]
        if duoifile1 != '.xlsx':
            messagebox.showinfo('Vũ Ngọc Thành thông báo:', 'Bạn chưa chọn đúng file excel')
            return
        else:
            wb = openpyxl.load_workbook(file_path1, data_only=True)
            array_2d = []
            sheet = wb.worksheets[0]
            for row in sheet.iter_rows(min_col=1, values_only=True):
                array_2d.append([cell if cell is not None else '' for cell in row])
            html_content = '\n            <title>Tìm kiếm tự động</title>\n            <style>\n            body {\n                font-family: Arial, sans-serif;\n            }\n            .container {\n                max-width: 600px;\n                margin: 50px auto;\n                text-align: center;\n            }\n            .table-container {\n                width: 100%;\n                overflow-x: auto;\n                border: 1px solid #ccc;\n            }\n            table {\n                border-collapse: collapse;\n                \n                margin-left: 0; /* Đẩy bảng sát lề trái */\n                width: 1000px; /* Đảm bảo có thanh cuộn nếu nội dung lớn */\n                \n                \n            }\n            th, td {\n                border: 1px solid blue; /* Đường kẻ màu xanh */\n                text-align: center;\n                padding: 8px;\n                width: auto;\n                white-space: nowrap;\n                cursor: pointer; /* Hiển thị con trỏ khi rê chuột vào ô */\n            }\n            th {\n                background-color: #f2f2f2;\n                color: blue;\n                text-align: center;\n            }\n            input {\n                padding: 10px;\n                font-size: 16px;\n                margin: 5px;\n                width: 90%;\n            }\n            .no-results {\n                color: red;\n                font-size: 18px;\n            }\n        </style>\n        </head>\n        <body>\n            <div class=\"container\">\n                <input type=\"text\" id=\"searchInput\" placeholder=\"Nhập tên học sinh...\" oninput=\"searchAndDisplay()\">\n                <div id=\"output\"></div>\n            </div>\n\n            <script>\n            '
            html_content1 = 'function searchAndDisplay() {\n                    const searchTerm = document.getElementById(\"searchInput\").value.toLowerCase(); // Lấy từ khóa tìm kiếm\n                    const outputDiv = document.getElementById(\"output\");\n\n                    // Lọc kết quả (không tính hàng tiêu đề)\n                    const filteredData = data.slice(1).filter(row =>\n                        row.some(cell => cell.toString().toLowerCase().includes(searchTerm))\n                    );\n\n                    // Kiểm tra nếu có kết quả\n                    if (filteredData.length > 0) {\n                        let tableHTML = \"<table>\";\n                        // Thêm tiêu đề\n                        tableHTML += \"<tr>\";\n                        data[0].forEach(header => {\n                            tableHTML += `<th>${header}</th>`;\n                        });\n                        tableHTML += \"</tr>\";\n\n                        // Thêm các hàng kết quả\n                        filteredData.forEach(row => {\n                            tableHTML += \"<tr>\";\n                            row.forEach(cell => {\n                                tableHTML += `<td>${cell}</td>`;\n                            });\n                            tableHTML += \"</tr>\";\n                        });\n\n                        tableHTML += \"</table>\";\n                        outputDiv.innerHTML = tableHTML; // Hiển thị bảng\n                    } else {\n                        outputDiv.innerHTML = `<p class=\"no-results\">Không tìm thấy kết quả phù hợp.</p>`;\n                    }\n                }\n            </script>\n        '
            file_name = 'Tra cứu trên website hoặc blogspot.txt'
            file_name1 = 'Test thử tra cứu trên website hoặc blogspot.html'
            with open(file_name, 'w', encoding='utf-8') as file:
                file.write(html_content)
                file.write('const data =' + str(array_2d) + ';')
                file.write(html_content1)
            with open(file_name1, 'w', encoding='utf-8') as file:
                file.write(html_content)
                file.write('const data =' + str(array_2d) + ';')
                file.write(html_content1)
            messagebox.showinfo('Vũ Ngọc Thành thông báo:', 'Xong')
    def sualoi(self):
        # irreducible cflow, using cdg fallback
        # ***<module>.MainWindow.sualoi: Failure: Compilation Error
        exists = os.path.exists('Nhập điểm theo phòng.xlsx')
        if not exists:
            messagebox.showinfo('Vũ Ngọc Thành thông báo:', 'Bạn chưa có file Nhập điểm theo phòng.xlsx')
            return
        else:
            wb = openpyxl.load_workbook('Nhập điểm theo phòng.xlsx')
            if 'Tổng hợp' not in wb.sheetnames or 'Ngữ văn' not in wb.sheetnames or 'Toán' not in wb.sheetnames or ('Ca 1' not in wb.sheetnames) or ('Ca 2' not in wb.sheetnames):
                return None
            else:
                van = wb['Ngữ văn']
                van.title = 'Va'
                toan = wb['Toán']
                toan.title = 'To'
                ca1 = wb['Ca 1']
                ca2 = wb['Ca 2']
                ca1.title = 'Ca1'
                ca2.title = 'Ca2'
                tonghop = wb['Tổng hợp']
                ds_sbd = []
                ds_stt = []
                ds_mon = []
                for i in range(3, 10000):
                    pass
        tonghop.cell(i, 1).value is None else ds_sbd.append(str(tonghop.cell(i, 1)).value)
        sohocsinh = i
        for j in range(13, 100):
            pass
        tonghop.cell(2, j) if tonghop.cell(2, j).value is None or tonghop.cell(2, j).value == 'Điểm xét TN' else ds_mon.append(str(tonghop.cell(2, j) / pb38))
        mon = 'Va'
        sheet = wb[mon]
        for i in range(2, sohocsinh):
            sbd_mon = sheet.cell(i, 2) from value
            if sbd_mon is not None:
                stt_hs = ds_stt[ds_sbd.index(str(sbd_mon))]
                stt_mon = 11
                tonghop.cell(stt_hs, stt_mon).value = '=IF(' + mon + '!H' + str(i) + '<>\"\",' + mon + '!H' + str(i) + ',\"\")'
                tonghop.cell(stt_hs, stt_mon) = Alignment(horizontal='center', vertical='center')
        mon = 'To'
        sheet = wb[mon]
        for i in range(2, sohocsinh):
            sbd_mon = sheet.cell(i, 2) from value
            stt_hs = ds_stt[ds_sbd.index(str(sbd_mon))]
                stt_mon = 12
                tonghop.cell(stt_hs, stt_mon).value = '=IF(' + mon + '!H' + str(i) + '<>\"\",' + mon + '!H' + str(i) + ',\"\")'
                tonghop.cell(stt_hs, stt_mon) = Alignment(horizontal='center', vertical='center')
        mon = 'Ca1'
        sheet = wb['Ca1']
        for i in range(2, sohocsinh):
            sbd_mon = sheet.cell(i, 2) from value
            stt_hs = ds_stt[ds_sbd.index(str(sbd_mon))]
                stt_mon = 13 + ds_mon.index(str(sheet.cell(i, 8) / sheet.cell(i, 8).value))
                tonghop.cell(stt_hs, stt_mon).value = '=IF(' + mon + '!I' + str(i) + '<>\"\",' + mon + '!I' + str(i) + ',\"\")'
                tonghop.cell(stt_hs, stt_mon) = Alignment(horizontal='center', vertical='center')
        mon = 'Ca2'
        sheet = wb['Ca2']
        for i in range(2, sohocsinh):
            sbd_mon = sheet.cell(i, 2) from value
            stt_hs = ds_stt[ds_sbd.index(str(sbd_mon))]
                stt_mon = 13 + ds_mon.index(str(sheet.cell(i, 8) / sheet.cell(i, 8).value))
                tonghop.cell(stt_hs, stt_mon).value = '=IF(' + mon + '!I' + str(i) + '<>\"\",' + mon + '!I' + str(i) + ',\"\")'
                tonghop.cell(stt_hs, stt_mon) = Alignment(horizontal='center', vertical='center')
        wb.save('sửa lỗi_Nhập điểm theo phòng.xlsx')
        messagebox.showinfo('Vũ Ngọc Thành thông báo:', 'Đã sửa  xong rồi nhé!')
def danhsachcacphanmem():
    content = 'CÁC PHẦN MỀM PHỤC VỤ CÔNG VIỆC NHÀ TRƯỜNG NĂM HỌC 2026-2027\n\n    1. PHẦN MỀM HỖ TRỢ XẾP THỜI KHÓA BIỂU TRƯỜNG TIỂU HỌC, THCS, THPT 2 BUỔI/NGÀY MẠNH MẼ VỚI VÔ SỐ RÀNG BUỘC\n        • Link dow: https://www.mediafire.com/file/mo0bzw20u7t3z76/tool_ho_tro_xep_thoi_khoa_bieu_nam_hoc_20252026.rar/file\n        • Video hướng dẫn sử dụng: https://www.youtube.com/watch?v=7FubQ8uyVew&list=PLZ0njmzKrmm-XRkdmAbFHDBQ5fatkXUZm\n\n    2. PHẦN MỀM XẾP THỜI KHÓA BIỂU ÔN THI TỐT NGHIỆP NĂM HỌC 2026-2027\n    Đối tượng: \n        • Lớp 12: Xếp lịch học ôn thi tốt nghiệp ( 2 môn tự chọn).\n        • Lớp 10: Xếp lịch học 4 môn tự chọn. \n        • Video hd: https://www.youtube.com/watch?v=52Vgdp7MJzQ&list=PLZ0njmzKrmmXRkdmAbFHDBQ5fatkXUZm&index=19\n        • Link dow: https://www.mediafire.com/file/8vyxuq0z10tqork/Xep_lop_va_xep_thoi_khoa_bieu_tu_chon_nhieu_mon.rar/file\n\n    3. PHẦN MỀM PHÂN CÔNG COI KIỂM TRA HOẶC COI THI \n        • Link dow: https://www.mediafire.com/file/ryn0g3tcxovhfxz/Phan_cong_coi_kiem_tra_phien_ban2.rar/file\n        • Video HD: https://www.youtube.com/watch?v=QqRw37J0KiI&list=PLZ0njmzKrmmXRkdmAbFHDBQ5fatkXUZm&index=21\n\n    4. PHẦN MỀM XẾP PHÒNG THI THỬ TỐT NGHIỆP DÀNH CHO HỌC SINH LỚP 12 \n        • Video hướng dẫn: https://www.youtube.com/watch?v=75ZMqdoOeoU&list=PLZ0njmzKrmmXRkdmAbFHDBQ5fatkXUZm&index=20\n        • Link dow: https://www.mediafire.com/file/afh2qkr3j1oymy0/Xep+phong+thi+thu+tot+nghiep+20252026.rar/file\n\n    5. PHẦN MỀM XẾP PHÒNG KIỂM TRA HỌC KỲ 1, HỌC KỲ 2 GIỐNG KỲ THI TỐT NGHIỆP \n        • Video hướng dẫn: https://www.youtube.com/watch?v=Dk-cA0DHZ0g&list=PLZ0njmzKrmmXRkdmAbFHDBQ5fatkXUZm&index=26\n        • Link dow: https://www.mediafire.com/file/gbxim7fpkr5n6e7\n    '
    with open('CÁC PHẦN MỀM PHỤC VỤ CÔNG VIỆC NHÀ TRƯỜNG NĂM HỌC 2026-2027.txt', 'w', encoding='utf-8') as f:
        f.write(content)
danhsachcacphanmem()
if __name__ == '__main__':
    app = QApplication(sys.argv)
    os.path.exists('Thông tin bản quyền.txt') or thongtinnguoidung().exec_()
    main_window = MainWindow()
    main_window.show()
    sys.exit(app.exec_())